import datetime
from datetime import date as dt_date, time as dtime, timedelta
import io
from io import StringIO
import json
import math
import os
import random
import re
import time
from collections import defaultdict
from django.db.models import Prefetch, Q
# 2. Third-party library imports (Tashqi kutubxonalar)
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle
from reportlab.pdfgen import canvas

# 3. Django core imports (Django freymvorkining o'ziga tegishli modullar)
from django.conf import settings
from django.contrib import messages
from django.contrib.admin.views.decorators import staff_member_required
from django.contrib.auth.decorators import login_required
from django.contrib.auth.models import User
from django.core import management
from django.core.mail import send_mail
from django.db import connection, transaction
from django.http import HttpResponse, JsonResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.utils.dateparse import parse_date

# 4. Local application imports (Sizning loyihangizga xos model va formalar)
from ..forms import StudentForm, StudentImportForm, TeacherForm, TeacherImportForm, SubjectForm
from ..models import (
    Attendance, Course, CourseGroup, Grade, Group, GroupSchedule,
    LANGUAGE_CHOICES, Room, Student, Subject, Teacher, DailyGrade
)



# ─────────────────────────────────────────
# KONSTANTALAR
# ─────────────────────────────────────────
PARA_TIMES = [
    (dtime(8, 30),  dtime(9, 50)),
    (dtime(10, 0),  dtime(11, 20)),
    (dtime(12, 0),  dtime(13, 20)),
    (dtime(13, 30), dtime(14, 50)),
    (dtime(15, 0),  dtime(16, 20)),
    (dtime(16, 30), dtime(17, 50)),
]

WEEKDAYS = {0: 'Dushanba', 1: 'Seshanba', 2: 'Chorshanba',
            3: 'Payshanba', 4: 'Juma', 5: 'Shanba', 6: 'Yakshanba'}

WEEKDAY_NAMES = {
    0: 'Dushanba', 1: 'Seshanba', 2: 'Chorshanba',
    3: 'Payshanba', 4: 'Juma', 5: 'Shanba'
}

WEEKDAY_OPTIONS = [
    (0, 'Dushanba'), (1, 'Seshanba'), (2, 'Chorshanba'),
    (3, 'Payshanba'), (4, 'Juma'), (5, 'Shanba'),
]

PARA_TIMES_WEEKLY = [
    (s.strftime("%H:%M"), e.strftime("%H:%M"))
    for s, e in PARA_TIMES
]

WEEKDAY_LIST = ["Dushanba", "Seshanba", "Chorshanba", "Payshanba", "Juma", "Shanba"]

GROUP_COLORS = [
    "D6E4BC", "B8D4E8", "FCE4A8", "E8C8D4",
    "CCE8CC", "FFD8B0", "D8D0E8", "E8E8C8",
    "BCE4E4", "FFC8C8", "D4E4F4", "E4D4BC",
    "C8D8F4", "F4D4C8", "D4F4D4", "F4F4C8",
]

VALID_PARA_PAIRS = [
    (0, 1),  # 1-blok: 08:30 - 11:20
    (2, 3),  # 2-blok: 12:00 - 14:50
    (4, 5),  # 3-blok: 15:00 - 17:50
]

# ── Guruh hajmi chegaralari (konflikt-hal qiluvchi barcha avtomatik
# funksiyalarda BIR XIL qo'llanadi — talaba ko'chirish/almashtirish/
# majburiy chiqarishda guruh bu chegaradan tashqariga chiqmasligi kerak) ──
MIN_GROUP_SIZE = 8   # guruhda bundan kam talaba qolishiga yo'l qo'yilmaydi
MAX_GROUP_SIZE = 18  # guruhga bundan ortiq talaba qo'shilishiga yo'l qo'yilmaydi

# ─────────────────────────────────────────
# YORDAMCHI FUNKSIYALAR
# ─────────────────────────────────────────
def is_admin(user):
    return user.is_superuser


def stats_api(request):
    return JsonResponse({
        'lessons': Course.objects.count(),
        'teachers': Teacher.objects.count(),
        'students': Student.objects.count(),
        'rooms': Room.objects.count(),
    })

def is_teacher(user):
    return hasattr(user, 'teacher')

def is_student(user):
    return hasattr(user, 'student')

def sync_group_language(group):
    """
    Guruhning `language` maydonini shu guruhdagi TALABALARNING haqiqiy
    ta'lim tiliga qarab avtomatik yangilaydi:
      - guruhda faqat 'uz' talabalar bo'lsa       -> 'uz'
      - guruhda faqat 'ru' talabalar bo'lsa       -> 'ru'
      - guruhda 'uz' VA 'ru' talabalar aralash    -> 'uz-ru'
      - boshqa tillar (masalan faqat 'qq' yoki 'en') bo'lsa, o'sha til(lar)
        alifbo tartibida '-' bilan qo'shib qo'yiladi (masalan 'qq' yoki 'en-uz')
      - guruhda hali talaba bo'lmasa -> hech narsa o'zgartirilmaydi

    Talabalar guruhga qo'shilgan/olib tashlangan/almashtirilgan HAR BIR
    joydan keyin albatta shu funksiya chaqirilishi kerak — shunda guruh nomi
    (tili) doim tarkibga mos bo'lib turadi (masalan bitta rus talaba qo'shilsa,
    sof o'zbek guruh avtomatik 'uz-ru' ga aylanadi).
    """
    langs = set(
        group.students.exclude(language__isnull=True)
        .exclude(language='')
        .values_list('language', flat=True)
    )
    if not langs:
        return

    if langs == {'uz'}:
        new_lang = 'uz'
    elif langs == {'ru'}:
        new_lang = 'ru'
    elif langs == {'uz', 'ru'}:
        new_lang = 'uz-ru'
    else:
        new_lang = '-'.join(sorted(langs))

    if group.language != new_lang:
        group.language = new_lang
        group.save(update_fields=['language'])