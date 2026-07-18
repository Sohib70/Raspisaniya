import datetime
from datetime import date as dt_date, time as dtime, timedelta
import io
from io import StringIO
import json
import math
import os
import random
import re
import time
from collections import defaultdict
from django.db.models import Prefetch
# 2. Third-party library imports (Tashqi kutubxonalar)
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle
from reportlab.pdfgen import canvas

# 3. Django core imports (Django freymvorkining o'ziga tegishli modullar)
from django.conf import settings
from django.contrib import messages
from django.contrib.admin.views.decorators import staff_member_required
from django.contrib.auth.decorators import login_required
from django.contrib.auth.models import User
from django.core import management
from django.core.mail import send_mail
from django.db import connection, transaction
from django.http import HttpResponse, JsonResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.utils.dateparse import parse_date

# 4. Local application imports (Sizning loyihangizga xos model va formalar)
from .forms import StudentForm, StudentImportForm, TeacherForm, TeacherImportForm, SubjectForm
from .models import (
    Attendance, Course, CourseGroup, Grade, Group, GroupSchedule,
    LANGUAGE_CHOICES, Room, Student, Subject, Teacher, DailyGrade
)



# ─────────────────────────────────────────
# KONSTANTALAR
# ─────────────────────────────────────────
PARA_TIMES = [
    (dtime(8, 30),  dtime(9, 50)),
    (dtime(10, 0),  dtime(11, 20)),
    (dtime(12, 0),  dtime(13, 20)),
    (dtime(13, 30), dtime(14, 50)),
    (dtime(15, 0),  dtime(16, 20)),
    (dtime(16, 30), dtime(17, 50)),
]

WEEKDAYS = {0: 'Dushanba', 1: 'Seshanba', 2: 'Chorshanba',
            3: 'Payshanba', 4: 'Juma', 5: 'Shanba', 6: 'Yakshanba'}

WEEKDAY_NAMES = {
    0: 'Dushanba', 1: 'Seshanba', 2: 'Chorshanba',
    3: 'Payshanba', 4: 'Juma', 5: 'Shanba'
}

WEEKDAY_OPTIONS = [
    (0, 'Dushanba'), (1, 'Seshanba'), (2, 'Chorshanba'),
    (3, 'Payshanba'), (4, 'Juma'), (5, 'Shanba'),
]

PARA_TIMES_WEEKLY = [
    (s.strftime("%H:%M"), e.strftime("%H:%M"))
    for s, e in PARA_TIMES
]

WEEKDAY_LIST = ["Dushanba", "Seshanba", "Chorshanba", "Payshanba", "Juma", "Shanba"]

GROUP_COLORS = [
    "D6E4BC", "B8D4E8", "FCE4A8", "E8C8D4",
    "CCE8CC", "FFD8B0", "D8D0E8", "E8E8C8",
    "BCE4E4", "FFC8C8", "D4E4F4", "E4D4BC",
    "C8D8F4", "F4D4C8", "D4F4D4", "F4F4C8",
]

VALID_PARA_PAIRS = [
    (0, 1),  # 1-blok: 08:30 - 11:20
    (2, 3),  # 2-blok: 12:00 - 14:50
    (4, 5),  # 3-blok: 15:00 - 17:50
]

# ── Guruh hajmi chegaralari (konflikt-hal qiluvchi barcha avtomatik
# funksiyalarda BIR XIL qo'llanadi — talaba ko'chirish/almashtirish/
# majburiy chiqarishda guruh bu chegaradan tashqariga chiqmasligi kerak) ──
MIN_GROUP_SIZE = 8   # guruhda bundan kam talaba qolishiga yo'l qo'yilmaydi
MAX_GROUP_SIZE = 18  # guruhga bundan ortiq talaba qo'shilishiga yo'l qo'yilmaydi

# ─────────────────────────────────────────
# YORDAMCHI FUNKSIYALAR
# ─────────────────────────────────────────
def is_admin(user):
    return user.is_superuser


def stats_api(request):
    return JsonResponse({
        'lessons': Course.objects.count(),
        'teachers': Teacher.objects.count(),
        'students': Student.objects.count(),
        'rooms': Room.objects.count(),
    })

def is_teacher(user):
    return hasattr(user, 'teacher')

def is_student(user):
    return hasattr(user, 'student')

def split_into_groups(students, max_size=15, min_size=10):
    total = len(students)
    if total == 0:
        return []
    num_groups = (total + max_size - 1) // max_size
    base_size = total // num_groups
    remainder = total % num_groups
    groups = []
    start = 0
    for i in range(num_groups):
        size = base_size + (1 if i < remainder else 0)
        groups.append(students[start:start + size])
        start += size
    return groups

def get_lesson_dates(start_date, weekdays, total):
    result = []
    cur = start_date
    while len(result) < total:
        if cur.weekday() in weekdays:
            result.append(cur)
        cur += timedelta(days=1)
    return result


def get_expected_lessons_per_week(total_lessons):
    """
    `find_schedule_for_group` kurs turi (24/16/8 paralik) uchun HAR DOIM
    qat'iy shu haftalik dars (para) sonini qo'llaydi — bu son `total_lessons`
    dan kelib chiqadi, `lessons_per_week` maydonidagi qiymatdan qat'iy nazar:
      - 24 paralik (>=20)  -> 6 para/hafta (3 kun x 2 para)
      - 16 paralik (12-20) -> 4 para/hafta (2 kun x 2 para)
      - 8 paralik  (<12)   -> 2 para/hafta (1 kun x 2 para)

    Kurs yaratish/tahrirlash formalarida admin "Haftada necha marta"
    maydoniga BOSHQA son kiritib qo'ysa, tugash sanasi (`weeks_needed`)
    noto'g'ri hisoblanib, jadval bilan sana bir-biriga mos kelmay qolardi.
    Shu funksiya orqali kiritilgan qiymat tekshiriladi/tuzatiladi.
    """
    if total_lessons >= 20:
        return 6
    elif 12 <= total_lessons <= 20:
        return 4
    else:
        return 2


def sync_group_language(group):
    """
    Guruhning `language` maydonini shu guruhdagi TALABALARNING haqiqiy
    ta'lim tiliga qarab avtomatik yangilaydi:
      - guruhda faqat 'uz' talabalar bo'lsa       -> 'uz'
      - guruhda faqat 'ru' talabalar bo'lsa       -> 'ru'
      - guruhda 'uz' VA 'ru' talabalar aralash    -> 'uz-ru'
      - boshqa tillar (masalan faqat 'qq' yoki 'en') bo'lsa, o'sha til(lar)
        alifbo tartibida '-' bilan qo'shib qo'yiladi (masalan 'qq' yoki 'en-uz')
      - guruhda hali talaba bo'lmasa -> hech narsa o'zgartirilmaydi

    Talabalar guruhga qo'shilgan/olib tashlangan/almashtirilgan HAR BIR
    joydan keyin albatta shu funksiya chaqirilishi kerak — shunda guruh nomi
    (tili) doim tarkibga mos bo'lib turadi (masalan bitta rus talaba qo'shilsa,
    sof o'zbek guruh avtomatik 'uz-ru' ga aylanadi).
    """
    langs = set(
        group.students.exclude(language__isnull=True)
        .exclude(language='')
        .values_list('language', flat=True)
    )
    if not langs:
        return

    if langs == {'uz'}:
        new_lang = 'uz'
    elif langs == {'ru'}:
        new_lang = 'ru'
    elif langs == {'uz', 'ru'}:
        new_lang = 'uz-ru'
    else:
        new_lang = '-'.join(sorted(langs))

    if group.language != new_lang:
        group.language = new_lang
        group.save(update_fields=['language'])


def parse_group_langs(lang_value):
    """
    Guruhning `language` maydonidagi tillarni to'plam qilib ajratib beradi.
    Masalan: 'uz' -> {'uz'}; 'ru' -> {'ru'}; 'uz-ru' -> {'uz', 'ru'}.
    Guruhlarni til bo'yicha solishtirishda (masalan parallel-swap uchun nomzod
    qidirishda) aniq matn tengligi ('uz' == 'uz-ru' -> False) o'rniga shu
    to'plamlar KESISHMASINI tekshirish kerak — shunda sof 'uz' guruh bilan
    aralash 'uz-ru' guruh ham bir-biriga mos nomzod sifatida ko'riladi.
    """
    if not lang_value:
        return set()
    return {part for part in lang_value.split('-') if part}


def get_student_group_conflict(student, target_group, exclude_group=None):
    """
    Talabaning `target_group`ga qo'shilishi, uning boshqa guruhlardagi mavjud
    dars vaqtlari bilan to'qnashib qolmasligini tekshiradi. Agar target_group
    hali jadvallanmagan bo'lsa (is_scheduled=False), tekshirishning hojati
    yo'q — u holda hozircha vaqt yo'q. To'qnashuv topilsa (sana, vaqt) juftini,
    aks holda None qaytaradi.
    """
    if not target_group.is_scheduled:
        return None
    target_times = set(
        GroupSchedule.objects.filter(group=target_group).values_list('date', 'start_time')
    )
    if not target_times:
        return None
    busy_qs = GroupSchedule.objects.filter(group__students=student)
    if exclude_group is not None:
        busy_qs = busy_qs.exclude(group=exclude_group)
    student_busy_times = set(busy_qs.values_list('date', 'start_time'))
    conflict_times = target_times & student_busy_times
    if conflict_times:
        return sorted(conflict_times)[0]
    return None


def get_teacher_group_conflict(teacher, target_group, exclude_group=None):
    """
    O'qituvchining `target_group`ga biriktirilishi (yoki target_group
    vaqtlarining o'zgarishi) o'qituvchining boshqa guruhlardagi mavjud
    dars vaqtlari bilan to'qnashib qolmasligini tekshiradi. To'qnashuv
    topilsa (sana, vaqt) juftini, aks holda None qaytaradi.
    """
    if not teacher or not target_group.is_scheduled:
        return None
    target_times = set(
        GroupSchedule.objects.filter(group=target_group).values_list('date', 'start_time')
    )
    if not target_times:
        return None
    busy_qs = GroupSchedule.objects.filter(group__teacher=teacher)
    if exclude_group is not None:
        busy_qs = busy_qs.exclude(group=exclude_group)
    teacher_busy_times = set(busy_qs.values_list('date', 'start_time'))
    conflict_times = target_times & teacher_busy_times
    if conflict_times:
        return sorted(conflict_times)[0]
    return None


def apply_lesson_time_change(sched, new_date_val, new_time_val, apply_to_future=False):
    """
    Bitta darsning (sched) vaqtini o'zgartiradi. `apply_to_future=True`
    bo'lsa VA sana o'zgarmagan (faqat vaqt o'zgargan) bo'lsa, xuddi shu
    guruhning KEYINGI barcha haftalaridagi bir xil hafta kuni + bir xil
    eski vaqtdagi darslari ham AVTOMATIK shu yangi vaqtga ko'chiriladi.

    MUHIM: agar SANA ham o'zgargan bo'lsa (masalan bayram sababli bitta
    darsni boshqa kunga ko'chirish), bu — istisno holat hisoblanadi va
    kaskad qilinmaydi, faqat o'sha bitta dars o'zgaradi. Kaskad faqat
    "doimiy vaqt o'zgarishi" (masalan "bugundan boshlab har doim soat
    14:00 da bo'ladi") uchun mo'ljallangan.

    Har bir nishon dars alohida to'qnashuv tekshiruvidan o'tadi — band
    bo'lgan sanalar o'tkazib yuboriladi (o'zgartirilmaydi), qolganlari
    yangilanadi.

    Qaytaradi: (updated_count, skipped_dates_list)
    """
    date_changed = new_date_val != sched.date
    old_time_val = sched.start_time
    old_weekday = sched.date.weekday()

    if apply_to_future and not date_changed:
        candidates = GroupSchedule.objects.filter(
            group=sched.group,
            date__gte=sched.date,
            start_time=old_time_val,
        )
        targets = [s for s in candidates if s.date.weekday() == old_weekday]
        if sched not in targets:
            targets.append(sched)
    else:
        targets = [sched]

    teacher_id = sched.group.teacher_id
    student_ids = list(sched.group.students.values_list('id', flat=True))

    updated = 0
    skipped_dates = []

    for t in targets:
        target_date = new_date_val if t.pk == sched.pk else t.date

        if teacher_id and GroupSchedule.objects.filter(
            date=target_date, start_time=new_time_val, group__teacher_id=teacher_id,
        ).exclude(pk=t.pk).exists():
            skipped_dates.append(target_date)
            continue

        if student_ids and GroupSchedule.objects.filter(
            date=target_date, start_time=new_time_val, group__students__id__in=student_ids,
        ).exclude(pk=t.pk).exists():
            skipped_dates.append(target_date)
            continue

        t.date = target_date
        t.start_time = new_time_val
        t.save(update_fields=['date', 'start_time'])
        updated += 1

    return updated, skipped_dates


def find_schedule_for_group(
        start_date, end_date, total_lessons, lessons_per_week,
        teacher=None, students=None, group_number=1,
        include_saturday=False,
        same_subject_busy=None
):
    from collections import defaultdict
    from datetime import timedelta

    if students is None:
        students = []
    if same_subject_busy is None:
        same_subject_busy = set()

    student_ids = [s.id for s in students]
    student_id_set = set(student_ids)
    teacher_id = teacher.id if teacher else None
    max_wd = 5 if include_saturday else 4
    available_wds = list(range(0, max_wd + 1))

    import itertools

    if total_lessons >= 20:
        # 24 para: Faqat Dush, Chor, Jum
        preferred = [(0, 2, 4)]
        days_needed = 3
    elif 12 <= total_lessons <= 20:
        # 16 para: Faqat Sesh, Pay
        preferred = [(1, 3)]
        days_needed = 2
    else:
        # 8 para: FAQAT bitta kun/hafta, 2 para/hafta (haftada 1 kun 2 para).
        # MUHIM: avvalgi versiya barcha bo'sh kunlarni BITTA pattern'ga yig'ib
        # olardi — bu bitta guruhning bitta haftasiga bir nechta kun (masalan
        # Sesh+Pay) sig'dirib, "haftada 1 kun 2 para" qoidasini buzardi va
        # darslarni kerakidan tezroq (kamroq haftaga) joylashtirib yuborardi.
        # Endi FAQAT bitta kun tanlanadi (days_needed=1). Lekin barcha
        # guruhlar bir xil kunga (odatda Dushanbaga) to'planib qolmasligi
        # uchun boshlang'ich kun guruh raqamiga (group_number) qarab
        # aylantiriladi — 1-guruh Dushanbadan, 2-guruh Seshanbadan va h.k.
        # boshlab qidiradi, shunda guruhlar tabiiy ravishda haftaning turli
        # kunlariga tarqaladi.
        start_offset = (group_number - 1) % len(available_wds)
        rotated_wds = available_wds[start_offset:] + available_wds[:start_offset]
        preferred = [(wd,) for wd in rotated_wds]
        days_needed = 1

    # 2. Qat'iy rejim: Faqat mavjud bo'lgan (start va end date oralig'iga tushadigan)
    # kunlarni filtrlash
    candidate_wd_sets = [c for c in preferred if all(wd in available_wds for wd in c)]

    week_monday = start_date - timedelta(days=start_date.weekday())

    def get_hard_busy(date):
        busy = set()
        if teacher_id:
            for sc in GroupSchedule.objects.filter(
                    date=date, group__teacher_id=teacher_id
            ).select_related('group'):
                st = sc.start_time or sc.group.start_time
                if st:
                    for i, (ps, _) in enumerate(PARA_TIMES):
                        if ps == st:
                            busy.add(i)
                else:
                    busy.update(range(len(PARA_TIMES)))

        if student_ids:
            for sc in GroupSchedule.objects.filter(
                    date=date, group__students__id__in=student_ids
            ).select_related('group').distinct():
                st = sc.start_time or sc.group.start_time
                if st:
                    for i, (ps, _) in enumerate(PARA_TIMES):
                        if ps == st:
                            busy.add(i)
                else:
                    busy.update(range(len(PARA_TIMES)))
        return busy

    def get_subject_busy_paras(date):
        # MUHIM: bu yerda ENDI binary (band/bo'sh) emas, balki HAR BIR parada
        # bitta fanning nechta guruhi allaqachon joylashganini SANAYMIZ.
        # Shu sonlar orqali eng "bo'sh" (kam guruh joylashgan) parani tanlab,
        # guruhlarni bir-biriga teng taqsimlashga erishamiz.
        counts = defaultdict(int)
        for (bd, bt) in same_subject_busy:
            if bd == date:
                for i, (ps, _) in enumerate(PARA_TIMES):
                    if ps == bt:
                        counts[i] += 1
        return counts

    # ── TUZATILGAN: Birinchi haftadan pattern qidirish mantiqi ──
    def find_best_pair(date):
        hard_busy = get_hard_busy(date)
        subject_busy_counts = get_subject_busy_paras(date)
        candidates = []

        for p1, p2 in VALID_PARA_PAIRS:
            # MUHIM: Ustoz yoki TALABA haqiqiy to'qnashuvi endi "ball" emas,
            # balki QAT'IY TAQIQ. Agar shu para juftlikda ustoz yoki talaba
            # boshqa darsga band bo'lsa, bu juftlik UMUMAN ko'rib chiqilmaydi —
            # aks holda bitta talaba bir vaqtda bir necha guruhga tushib qolardi.
            student_conflicts = sum(1 for p in (p1, p2) if p in hard_busy)
            if student_conflicts > 0:
                continue

            # Faqat to'qnashuvsiz paralar orasida — bitta fanning kamroq band
            # bo'lgan parasini tanlaymiz (guruhlarni teng taqsimlash uchun)
            subj_conflicts = subject_busy_counts[p1] + subject_busy_counts[p2]
            candidates.append((subj_conflicts, p1, p2))

        if not candidates:
            return None

        candidates.sort(key=lambda x: x[0])
        best = candidates[0]
        return (best[1], best[2], best[0])

    def get_busy_detailed(date):
        busy = defaultdict(list)
        if teacher_id:
            for sc in GroupSchedule.objects.filter(
                    date=date, group__teacher_id=teacher_id
            ).select_related('group__course__subject', 'group__teacher'):
                st = sc.start_time or sc.group.start_time
                idxs = list(range(len(PARA_TIMES))) if not st else [
                    i for i, (ps, _) in enumerate(PARA_TIMES) if ps == st
                ]
                for i in idxs:
                    busy[i].append({
                        'type': 'teacher', 'group': sc.group,
                        'subject': sc.group.course.subject,
                        'busy_students': []
                    })
        if student_ids:
            for sc in GroupSchedule.objects.filter(
                    date=date, group__students__id__in=student_ids
            ).select_related('group__course__subject', 'group__teacher') \
                    .prefetch_related('group__students').distinct():
                st = sc.start_time or sc.group.start_time
                idxs = list(range(len(PARA_TIMES))) if not st else [
                    i for i, (ps, _) in enumerate(PARA_TIMES) if ps == st
                ]
                busy_sts = [s for s in sc.group.students.all()
                            if s.id in student_id_set]
                for i in idxs:
                    busy[i].append({
                        'type': 'student', 'group': sc.group,
                        'subject': sc.group.course.subject,
                        'busy_students': busy_sts
                    })
        return busy

    # ── FAQAT start_date haftasida pattern qidirish — BARCHA kombinatsiyalar ──
    pattern = []

    for wd_set in candidate_wd_sets:
        trial_pattern = []
        for wd in wd_set:
            d = week_monday + timedelta(days=wd)
            if d < start_date or d > end_date:
                continue
            pair = find_best_pair(d)
            if pair is not None:
                trial_pattern.append((wd, pair[0], pair[1]))

        if len(trial_pattern) >= days_needed:
            pattern = trial_pattern
            break
        # Bu kombinatsiya ishlamadi — keyingisini sinaymiz (natijani tashlab yubormaymiz,
        # faqat eng oxirida hech biri ishlamasa, xabar uchun saqlab qo'yamiz)

    # ── C variant: HECH QANDAY kombinatsiya ishlamadi → xato ───────
    if len(pattern) < days_needed:
        conflict_info = []
        for wd in available_wds:
            d = week_monday + timedelta(days=wd)
            if d < start_date or d > end_date:
                continue
            bd = get_busy_detailed(d)
            for pi, occs in bd.items():
                for occ in occs:
                    conflict_info.append({
                        'date': d,
                        'para_index': pi,
                        'para_time': PARA_TIMES[pi],
                        'type': occ['type'],
                        'group': occ['group'],
                        'subject': occ['subject'],
                        'busy_students': occ['busy_students'],
                    })
        find_schedule_for_group._last_conflict_info = conflict_info
        find_schedule_for_group._last_missing = total_lessons
        find_schedule_for_group._last_no_slot_in_week = True
        return []
    find_schedule_for_group._last_no_slot_in_week = False
    result = []
    cur_monday = week_monday
    while len(result) < total_lessons:
        if cur_monday > end_date + timedelta(weeks=24):
            break

        for (wd, p1, p2) in pattern:
            if len(result) >= total_lessons:
                break

            d = cur_monday + timedelta(days=wd)

            # Sanani tekshiramiz
            if d < start_date or d > end_date:
                continue

            # ✅ YANGI TEKSHIRUV: Agar o'sha kunda ustoz yoki talaba band bo'lsa,
            # darsni o'sha haftada emas, keyingi haftada davom ettirish kerakmi?
            # Yo'q, biz "qat'iy rejim"damiz, shuning uchun darsni yozamiz
            # (chunki biz allaqachon to'qnashuvsiz kunni topganmiz).

            remaining = total_lessons - len(result)

            # Juft para (p1 va p2) mantiqi
            if remaining >= 2:
                result.append((d, PARA_TIMES[p1][0], PARA_TIMES[p1][1]))
                result.append((d, PARA_TIMES[p2][0], PARA_TIMES[p2][1]))
            else:
                # Agar 1 ta dars qolgan bo'lsa
                result.append((d, PARA_TIMES[p1][0], PARA_TIMES[p1][1]))

        cur_monday += timedelta(weeks=1)

    result.sort(key=lambda x: (x[0], x[1]))

    missing = max(0, total_lessons - len(result))
    conflict_info = []

    if missing > 0:
        chk = week_monday
        for _ in range(10):
            if chk > end_date + timedelta(weeks=12):
                break
            for (wd, p1, p2) in pattern:
                d = chk + timedelta(days=wd)
                if d < start_date:
                    continue
                bd = get_busy_detailed(d)
                for pi in (p1, p2):
                    if pi in bd:
                        for occ in bd[pi]:
                            conflict_info.append({
                                'date': d,
                                'para_index': pi,
                                'para_time': PARA_TIMES[pi],
                                'type': occ['type'],
                                'group': occ['group'],
                                'subject': occ['subject'],
                                'busy_students': occ['busy_students'],
                            })
            chk += timedelta(weeks=1)
            if len(conflict_info) >= 30:
                break

    find_schedule_for_group._last_conflict_info = conflict_info
    find_schedule_for_group._last_missing = missing
    return result


def _auto_resolve_via_cross_subject_swap(grp_a, conflicts, group_last_positions=None):
    """
    grp_a joylasha olmayapti — boshqa fanning joylashgan guruhi
    bilan VAQT almashish orqali joy ochadi.

    group_last_positions: dict ko'rinishida guruhlarning oxirgi band qilgan vaqti saqlanadi.
    Shunda guruh faqat o'zining eski vaqtiga qayta olmaydi (Ping-pong bo'lmaydi),
    lekin boshqa istalgan yangi paraga ko'chaveradi.
    """
    if not conflicts:
        return None

    if group_last_positions is None:
        group_last_positions = {}  # {group_id: (date, start_time_string)}

    course = grp_a.course
    start = course.start_date
    week_monday = start - timedelta(days=start.weekday())

    # Kurs darslar soniga qarab kunlarni aniqlash
    if course.total_lessons >= 20:
        needed_wds = [0, 2, 4]  # Dush, Chor, Jum
    elif course.total_lessons >= 12 and course.total_lessons <= 20:
        needed_wds = [1, 3]  # Sesh, Pay
    else:
        needed_wds = list(range(5))

    # conflicts dan qaysi kun/paralar band ekanini aniqlaymiz
    blocked = defaultdict(set)
    for c in conflicts:
        c_date = c.get('date')
        c_pi = c.get('para_index')
        if c_date and c_pi is not None:
            blocked[c_date].add(c_pi)

    for wd in needed_wds:
        d = week_monday + timedelta(days=wd)
        if d < start or d > course.end_date:
            continue
        if d not in blocked:
            continue

        blocked_paras = blocked[d]

        for pi in list(blocked_paras):
            # Shu kunda, shu parada turgan boshqa guruhlarni topamiz
            blocking_scheds = GroupSchedule.objects.filter(
                date=d,
                start_time=PARA_TIMES[pi][0],
                group__is_scheduled=True,
            ).exclude(
                group=grp_a
            ).select_related(
                'group__course__subject', 'group__teacher'
            ).prefetch_related('group__students')

            for b_sched in blocking_scheds:
                b_grp = b_sched.group
                b_teacher_id = b_grp.teacher_id
                b_student_ids = list(b_grp.students.values_list('id', flat=True))

                # b_grp ning juft dars parasini topamiz
                partner_pi = None
                for pp1, pp2 in VALID_PARA_PAIRS:
                    if pp1 == pi:
                        partner_pi = pp2
                        break
                    if pp2 == pi:
                        partner_pi = pp1
                        break

                # b_grp ni ko'chirish mumkin bo'lgan yangi juft parani qidiramiz
                for new_p1, new_p2 in VALID_PARA_PAIRS:
                    # Eski para bilan ustma-ust kelib qolmasin
                    if new_p1 == pi or new_p2 == pi:
                        continue
                    if partner_pi is not None and (new_p1 == partner_pi or new_p2 == partner_pi):
                        continue

                    # ── PING-PONG OLDINI OLISH: Faqat o'zining eski o'rniga qaytishni taqiqlaymiz ──
                    target_time_1 = PARA_TIMES[new_p1][0]
                    last_pos = group_last_positions.get(b_grp.id)
                    if last_pos and last_pos == (d, target_time_1):
                        # Agar guruh xuddi shu kunda, shu paraga qaytayotgan bo'lsa - rad etamiz
                        continue

                    # 1. Yangi parada o'qituvchi band emasmi?
                    if b_teacher_id:
                        t_busy = GroupSchedule.objects.filter(
                            date=d,
                            start_time__in=[PARA_TIMES[new_p1][0], PARA_TIMES[new_p2][0]],
                            group__teacher_id=b_teacher_id,
                        ).exclude(group=b_grp).exists()
                        if t_busy:
                            continue

                    # 2. Yangi parada talabalar band emasmi?
                    if b_student_ids:
                        s_busy = GroupSchedule.objects.filter(
                            date=d,
                            start_time__in=[PARA_TIMES[new_p1][0], PARA_TIMES[new_p2][0]],
                            group__students__id__in=b_student_ids,
                        ).exclude(group=b_grp).exists()
                        if s_busy:
                            continue

                    # ✅ Yangi vaqt topildi -> Ko'chiramiz
                    with transaction.atomic():
                        moved = False

                        # Eski joylashuvni xotiraga yozamiz (Qaytmasligi uchun qulflash)
                        current_time = PARA_TIMES[pi][0]
                        group_last_positions[b_grp.id] = (d, current_time)

                        # 1-parani ko'chirish
                        s1 = GroupSchedule.objects.filter(
                            date=d, group=b_grp,
                            start_time=current_time
                        ).first()
                        if s1:
                            s1.start_time = target_time_1
                            s1.save(update_fields=['start_time'])
                            moved = True

                        # 2-parani ko'chirish
                        if partner_pi is not None:
                            s2 = GroupSchedule.objects.filter(
                                date=d, group=b_grp,
                                start_time=PARA_TIMES[partner_pi][0]
                            ).first()
                            if s2:
                                s2.start_time = PARA_TIMES[new_p2][0]
                                s2.save(update_fields=['start_time'])

                    if moved:
                        # ✅ Tuzatildi: O'qituvchi ismi obyektning o'zini stringga o'girish orqali xavfsiz olindi
                        b_teacher_name = str(b_grp.teacher) if b_grp.teacher else "O'qituvchi biriktirilmagan"
                        return (
                            f"⚡ Fano'g'ri almashtirish (Cross-Subject): '{b_grp.course.subject}' fani "
                            f"'{b_grp.group_number}-guruh' ({b_teacher_name}) {d.strftime('%d.%m.%Y')} kunidagi "
                            f"{pi + 1}-paradan {new_p1 + 1}-paraga muvaffaqiyatli ko'chirildi."
                        )
    return None

def _brute_force_find_slot(grp_a):
    """
    grp_a uchun joy qidiradi. Barcha `is_scheduled=True` guruhlarni
    birin-ketin ko'rib chiqadi va ularni boshqa bo'sh joyga surishga harakat qiladi.
    """
    course = grp_a.course
    start = course.start_date
    end = course.end_date
    week_monday = start - timedelta(days=start.weekday())
    include_saturday = getattr(course, 'include_saturday', False)
    max_wd = 5 if include_saturday else 4

    if course.total_lessons >= 20:
        needed_wds = [wd for wd in (0, 2, 4) if wd <= max_wd]
    elif course.total_lessons >= 12:
        needed_wds = [wd for wd in (1, 3) if wd <= max_wd]
    else:
        needed_wds = list(range(max_wd + 1))

    grp_a_teacher_id = grp_a.teacher_id
    grp_a_student_ids = list(grp_a.students.values_list('id', flat=True))

    needed_slots = []  # [(date, para_idx), ...]
    for wd in needed_wds:
        d = week_monday + timedelta(days=wd)
        if d < start or d > end:
            continue
        for pi in range(len(PARA_TIMES)):
            blocked_by_teacher = False
            blocked_by_student = False

            if grp_a_teacher_id:
                blocked_by_teacher = GroupSchedule.objects.filter(
                    date=d,
                    start_time=PARA_TIMES[pi][0],
                    group__teacher_id=grp_a_teacher_id,
                ).exclude(group=grp_a).exists()

            if grp_a_student_ids and not blocked_by_teacher:
                blocked_by_student = GroupSchedule.objects.filter(
                    date=d,
                    start_time=PARA_TIMES[pi][0],
                    group__students__id__in=grp_a_student_ids,
                ).exclude(group=grp_a).exists()

            if blocked_by_teacher or blocked_by_student:
                needed_slots.append((d, pi))

    if not needed_slots:
        return None

    all_scheduled = list(
        CourseGroup.objects.filter(
            is_scheduled=True,
        ).exclude(
            pk=grp_a.pk,
        ).select_related(
            'course__subject', 'teacher'
        ).prefetch_related('students')
        .order_by('pk')
    )

    if not all_scheduled:
        return None

    visited = set()
    queue = list(all_scheduled)

    while queue:
        blocker = queue.pop(0)

        if blocker.pk in visited:
            continue
        visited.add(blocker.pk)

        b_teacher_id = blocker.teacher_id
        b_student_ids = list(blocker.students.values_list('id', flat=True))

        blocker_slots = []
        for wd in needed_wds:
            d = week_monday + timedelta(days=wd)
            if d < start or d > end:
                continue
            for sc in GroupSchedule.objects.filter(date=d, group=blocker):
                st = sc.start_time or blocker.start_time
                if not st:
                    continue
                for i, (ps, _) in enumerate(PARA_TIMES):
                    if ps == st:
                        is_problem = (d, i) in needed_slots
                        if is_problem:
                            blocker_slots.append((d, i, sc))

        if not blocker_slots:
            continue

        blocker_current_paras = defaultdict(set)
        for sc2 in GroupSchedule.objects.filter(
                date__gte=start, date__lte=end, group=blocker
        ):
            st2 = sc2.start_time or blocker.start_time
            if st2:
                for i, (ps, _) in enumerate(PARA_TIMES):
                    if ps == st2:
                        blocker_current_paras[sc2.date].add(i)

        for (prob_date, prob_pi, prob_sc) in blocker_slots:
            blk_own_paras = blocker_current_paras.get(prob_date, set())

            partner_pis = []
            for pp1, pp2 in VALID_PARA_PAIRS:
                if pp1 == prob_pi:
                    partner_pis.append(pp2)
                elif pp2 == prob_pi:
                    partner_pis.append(pp1)
            partner_pi = partner_pis[0] if partner_pis else None

            blk_others_busy = set()
            if b_teacher_id:
                for sc3 in GroupSchedule.objects.filter(
                        date=prob_date, group__teacher_id=b_teacher_id
                ).exclude(group=blocker):
                    st3 = sc3.start_time or sc3.group.start_time
                    if st3:
                        for i, (ps, _) in enumerate(PARA_TIMES):
                            if ps == st3:
                                blk_others_busy.add(i)
            if b_student_ids:
                for sc3 in GroupSchedule.objects.filter(
                        date=prob_date,
                        group__students__id__in=b_student_ids,
                ).exclude(group=blocker).distinct():
                    st3 = sc3.start_time or sc3.group.start_time
                    if st3:
                        for i, (ps, _) in enumerate(PARA_TIMES):
                            if ps == st3:
                                blk_others_busy.add(i)

            grp_a_busy_today = set()
            if grp_a_teacher_id:
                for sc3 in GroupSchedule.objects.filter(
                        date=prob_date, group__teacher_id=grp_a_teacher_id
                ).exclude(group=grp_a):
                    st3 = sc3.start_time or sc3.group.start_time
                    if st3:
                        for i, (ps, _) in enumerate(PARA_TIMES):
                            if ps == st3:
                                grp_a_busy_today.add(i)
            if grp_a_student_ids:
                for sc3 in GroupSchedule.objects.filter(
                        date=prob_date,
                        group__students__id__in=grp_a_student_ids,
                ).exclude(group=grp_a).distinct():
                    st3 = sc3.start_time or sc3.group.start_time
                    if st3:
                        for i, (ps, _) in enumerate(PARA_TIMES):
                            if ps == st3:
                                grp_a_busy_today.add(i)

            for new_p1, new_p2 in VALID_PARA_PAIRS:
                if new_p1 in blk_own_paras or new_p2 in blk_own_paras:
                    continue
                if new_p1 in blk_others_busy or new_p2 in blk_others_busy:
                    continue
                if new_p1 in grp_a_busy_today or new_p2 in grp_a_busy_today:
                    continue
                needed_pis_today = {pi for (d, pi) in needed_slots if d == prob_date}
                if new_p1 in needed_pis_today or new_p2 in needed_pis_today:
                    continue

                try:
                    with transaction.atomic():
                        moved_count = 0

                        s1 = GroupSchedule.objects.filter(
                            date=prob_date,
                            group=blocker,
                            start_time=PARA_TIMES[prob_pi][0],
                        ).first()
                        if s1:
                            s1.start_time = PARA_TIMES[new_p1][0]
                            s1.save(update_fields=['start_time'])
                            moved_count += 1

                        if partner_pi is not None:
                            s2 = GroupSchedule.objects.filter(
                                date=prob_date,
                                group=blocker,
                                start_time=PARA_TIMES[partner_pi][0],
                            ).first()
                            if s2:
                                s2.start_time = PARA_TIMES[new_p2][0]
                                s2.save(update_fields=['start_time'])
                                moved_count += 1
                            else:
                                raise ValueError("Juft para topilmadi")

                        if moved_count == 0:
                            raise ValueError("Ko'chiriladigan dars yo'q")

                except (ValueError, Exception):
                    continue

                # ✅ TUZATILDI: Muvaffaqiyatli xabar matni to'liq shakllantirildi
                old_t = PARA_TIMES[prob_pi][0].strftime('%H:%M')
                new_t = PARA_TIMES[new_p1][0].strftime('%H:%M')
                b_teacher_name = str(blocker.teacher) if blocker.teacher else "O'qituvchi biriktirilmagan"
                return (
                    f"⚡ Brute-force muvaffaqiyatli: '{blocker.course.subject}' fani "
                    f"'{blocker.group_number}-guruh' ({b_teacher_name}) {prob_date.strftime('%d.%m.%Y')} kunidagi "
                    f"soat {old_t} dars vaqti {new_t} ga surildi va '{grp_a.course.subject}' uchun joy ochildi."
                )

    return None

# ... avvalgi yordamchi funksiyalar (subject_swap, parallel_swap, cross_subject_swap, brute_force) ...


# ── 1. SHU YERGA YANGI MAJBURIY CHIQARISH FUNKSIYASINI JOYLASHTIRING ──
def _auto_resolve_by_force_student_eviction(grp_a, conflicts, same_course_groups):
    """
    Agar grp_a guruhiga dars qo'yishga ziddiyat keltirib chiqarayotgan talabalar soni
    juda kam bo'lsa (1 tadan 3 tagacha), ularni ushbu guruhdan majburlab chiqaradi va
    boshqa ziddiyatsiz parallel guruhlarga tarqatadi.
    Guruh hajmi chegaralari: MIN_GROUP_SIZE / MAX_GROUP_SIZE (yuqorida belgilangan).
    """
    if not conflicts:
        return None

    eviction_candidates = set()
    for c in conflicts:
        if c.get('type') == 'student':
            for st in c.get('busy_students', []):
                if grp_a.students.filter(pk=st.pk).exists():
                    eviction_candidates.add(st)

    if not eviction_candidates or len(eviction_candidates) > 3:
        return None

    other_groups = [g for g in same_course_groups if g.pk != grp_a.pk]
    if not other_groups:
        return None

    migrations = []
    temp_counts = {g.pk: g.students.count() for g in other_groups}

    for st in eviction_candidates:
        moved = False
        for target_grp in other_groups:
            if temp_counts[target_grp.pk] + 1 > MAX_GROUP_SIZE:
                continue

            if target_grp.is_scheduled:
                target_times = set(
                    GroupSchedule.objects.filter(group=target_grp).values_list('date', 'start_time')
                )
                student_busy = set(
                    GroupSchedule.objects.filter(group__students=st)
                    .exclude(group=grp_a)
                    .values_list('date', 'start_time')
                )
                if student_busy & target_times:
                    continue

            migrations.append((st, target_grp))
            temp_counts[target_grp.pk] += 1
            moved = True
            break

        if not moved:
            return None

    # Guruhda kamida MIN_GROUP_SIZE ta o'quvchi qolishi shart!
    if (grp_a.students.count() - len(eviction_candidates)) < MIN_GROUP_SIZE:
        return None

    with transaction.atomic():
        evicted_names = []
        affected_groups = {grp_a}
        for st, target_grp in migrations:
            grp_a.students.remove(st)
            target_grp.students.add(st)
            affected_groups.add(target_grp)
            evicted_names.append(f"{st.first_name} (-> {target_grp.group_number}-guruh)")

        for g in affected_groups:
            sync_group_language(g)

        evicted_str = ", ".join(evicted_names)
        return (
            f"🔄 Majburiy ko'chirish: '{grp_a}' guruhiga dars qo'yilishiga xalaqit berayotgan "
            f"kam sonli talabalar {evicted_str} boshqa guruhlarga surildi (guruh tarkibi 8 tadan kam bo'lib qolmadi)."
        )


def _try_dissolve_and_distribute_group(grp, same_course_groups):
    """
    Muammoli grp guruhini o'chirib, uning talabalarini boshqa guruhlarga tarqatadi.
    Faqat barcha talabalar muvaffaqiyatli joylashsa (sig'im <= MAX_GROUP_SIZE va
    dars vaqti to'qnashuvlarisiz), o'zgarishni bazada saqlaydi. Aks holda rad etadi (Rollback).
    """
    students_to_distribute = list(grp.students.all())
    other_groups = [g for g in same_course_groups if g.pk != grp.pk]

    if not other_groups:
        return False, "Tarqatish uchun boshqa parallel guruhlar yo'q."

    # Guruhlar sig'imi va talabalar ro'yxatini xotirada vaqtincha simulyatsiya qilamiz
    temp_assignments = {g.pk: list(g.students.all()) for g in other_groups}

    for student in students_to_distribute:
        distributed = False
        for target_grp in other_groups:
            current_students_in_target = temp_assignments[target_grp.pk]

            # 1. Yangi sig'im tekshiruvi
            if len(current_students_in_target) >= MAX_GROUP_SIZE:
                continue

            # 2. Agar maqsadli guruh allaqachon dars jadvaliga ega bo'lsa, talabaning vaqti mos keladimi?
            if target_grp.is_scheduled:
                target_times = set(
                    GroupSchedule.objects.filter(group=target_grp).values_list('date', 'start_time')
                )
                student_busy = set(
                    GroupSchedule.objects.filter(group__students=student)
                    .exclude(group=grp)  # joriy o'chirilayotgan guruh darslarini hisobga olmaymiz
                    .values_list('date', 'start_time')
                )
                # Agar talaba maqsadli guruh dars vaqtida band bo'lsa, bu guruhga qo'sha olmaymiz
                if student_busy & target_times:
                    continue

            # Shartlar bajarilsa, talabani vaqtincha shu guruhga yozamiz
            current_students_in_target.append(student)
            distributed = True
            break

        if not distributed:
            # Bitta talaba bo'lsa ham joylasha olmay qolsa, tarqatishni butunlay bekor qilamiz
            return False, f"Ba'zi talabalar guruh sig'imi (max {MAX_GROUP_SIZE}) yoki dars vaqti to'qnashuvi sababli boshqa guruhlarga sig'madi."

    # Hamma muvaffaqiyatli tarqaldi -> o'zgarishlarni bazada atomar saqlaymiz
    with transaction.atomic():
        for target_grp in other_groups:
            target_grp.students.set(temp_assignments[target_grp.pk])
            sync_group_language(target_grp)

        # Eski muammoli guruhni butunlay o'chirib tashlaymiz
        GroupSchedule.objects.filter(group=grp).delete()
        grp.delete()

    return True, f"'{grp}' guruhi dars jadvalida bo'sh joy topilmagani sababli tarqatib yuborildi. Talabalar qolgan guruhlarga muvaffaqiyatli qo'shildi."


def _auto_resolve_conflicts_by_subject_swap(grp_a, conflicts, already_moved_student_ids=None):
    """
    Parallel guruhlar orasida o'quvchilarni almashtiradi.
    Til (language) cheklovi mutlaqo olib tashlandi - o'zbek va rus guruh o'quvchilari o'rin almasha oladi!
    Guruh hajmi chegaralari: MIN_GROUP_SIZE / MAX_GROUP_SIZE (yuqorida belgilangan) —
    manba guruh (og) MIN_GROUP_SIZE dan kamayib qolmasligi, maqsadli guruh (cand) esa
    MAX_GROUP_SIZE dan oshib ketmasligi qat'iy tekshiriladi.
    """
    messages_out = []
    if not conflicts:
        return messages_out

    if already_moved_student_ids is None:
        already_moved_student_ids = set()

    grp_a_student_ids = set(grp_a.students.values_list('id', flat=True))

    conflict_map = defaultdict(set)
    for c in conflicts:
        if c.get('type') != 'student':
            continue
        for st in c.get('busy_students', []):
            if st.pk in already_moved_student_ids:
                continue
            conflict_map[(st, c['group'])].add((c['date'], c['para_time'][0]))

    if not conflict_map:
        return messages_out

    for (st, og), conflict_times in conflict_map.items():
        if st.pk in already_moved_student_ids:
            continue

        if not og.students.filter(pk=st.pk).exists():
            continue

        oc = og.course
        og_student_count = og.students.count()
        # MUHIM: `og`ning O'ZINING to'liq jadvali kerak — ret_st (qaytaruvchi
        # talaba) shu guruhga qo'shilganda haqiqiy to'qnashuv chiqmasligini
        # tekshirish uchun (avval bu o'rniga grp_a.teacher jadvali solishtirilardi —
        # bu NOTO'G'RI edi, chunki ret_st OG ga qo'shiladi, grp_a ga emas).
        og_times = set(
            GroupSchedule.objects.filter(group=og).values_list('date', 'start_time')
        )

        # MUHIM: `st` (ko'chiriladigan talaba)ning TO'LIQ band jadvali —
        # faqat grp_a bilan bog'liq to'qnashuvlar (conflict_times) emas.
        # Aks holda st boshqa, mutlaqo aloqasi bo'lmagan biror fanga
        # yangi guruhda (cand) tasodifan to'qnashib qolishi mumkin edi.
        st_full_busy = set(
            GroupSchedule.objects.filter(group__students=st)
            .exclude(group__in=[og, grp_a])
            .values_list('date', 'start_time')
        )

        # MUHIM: `course=oc` (aynan bitta Course yozuvi) o'rniga endi
        # `course__subject=oc.subject` — chunki bir xil fan turli Course
        # sifatida (masalan turli oqim/fakultet uchun) yaratilgan bo'lishi
        # mumkin. Avvalgi qidiruv bunday hollarda parallel guruhni umuman
        # topa olmasdi.
        candidates = CourseGroup.objects.filter(
            course__subject=oc.subject, is_scheduled=True
        ).exclude(pk=og.pk).select_related('teacher').prefetch_related(
            'students',
            Prefetch('schedule', to_attr='cached_schedules')
        )

        if not candidates.exists():
            continue

        # Hamma nomzod talabalarni yig'ish (Tilidan qat'iy nazar)
        all_candidate_student_ids = set()
        for cand in candidates:
            for ret_st in cand.students.all():
                if ret_st.id != st.pk:
                    all_candidate_student_ids.add(ret_st.id)

        student_busy_map = defaultdict(set)
        if all_candidate_student_ids:
            ret_busy_schedules = GroupSchedule.objects.filter(
                group__students__in=all_candidate_student_ids
            ).values('group__students__id', 'date', 'start_time')

            for sch in ret_busy_schedules:
                student_busy_map[sch['group__students__id']].add((sch['date'], sch['start_time']))

        for cand in candidates:
            cand_times = {(sch.date, sch.start_time) for sch in cand.cached_schedules}
            if conflict_times & cand_times:
                continue
            # YANGI TEKSHIRUV: st ning TO'LIQ jadvali cand bilan to'qnashmasin
            if st_full_busy & cand_times:
                continue

            cand_student_count = cand.students.count()

            # Nomzod talabalar ro'yxati (Til cheklovisiz)
            cand_students = [
                ret_st for ret_st in cand.students.all()
                if ret_st.id != st.pk
            ]

            safe_return = None
            if cand_students:
                for ret_st in cand_students:
                    ret_busy_times = student_busy_map[ret_st.id]
                    # TUZATILGAN: ret_st OG ga qo'shiladi — demak uning band
                    # vaqtlari OG ning O'ZINING jadvali bilan solishtirilishi kerak
                    conflict_for_ret = bool(ret_busy_times & og_times)

                    if not conflict_for_ret:
                        safe_return = ret_st
                        break

            # ── QAT'IY SHART 1: Agar safe_return (o'rniga keladigan) topilmasa
            # va original guruhda o'quvchi soni MIN_GROUP_SIZE dan kamayib qolsa,
            # ko'chirishga mutlaqo yo'l qo'ymaymiz ──
            if not safe_return and (og_student_count - 1) < MIN_GROUP_SIZE:
                continue

            # ── QAT'IY SHART 2 (YANGI): Agar safe_return topilmasa (bir tomonlama
            # ko'chirish bo'ladi), cand guruh MAX_GROUP_SIZE dan oshib ketmasligi
            # kerak. safe_return bo'lsa, cand hajmi o'zgarmaydi (bittasi chiqib,
            # bittasi kiradi) — shuning uchun bu holatda tekshirish shart emas.
            if not safe_return and cand_student_count + 1 > MAX_GROUP_SIZE:
                continue

            with transaction.atomic():
                og.students.remove(st)
                cand.students.add(st)

                log_msg = f"O'quvchi {st} '{og}' guruhidan '{cand}' guruhiga ko'chirildi."

                if safe_return:
                    cand.students.remove(safe_return)
                    og.students.add(safe_return)
                    log_msg += f" O'rniga '{safe_return}' qaytarildi (Swap)."
                else:
                    log_msg += f" Bir tomonlama ko'chirish bajarildi (Guruhda {og_student_count - 1} o'quvchi qoldi)."

                sync_group_language(og)
                sync_group_language(cand)

                messages_out.append(log_msg)

            already_moved_student_ids.add(st.pk)
            break

    return messages_out


def _auto_resolve_via_parallel_swap(grp_a):
    """
    Guruhdagi eng ziddiyatli talabani topib, uni parallel guruhga o'tkazadi.
    O'qituvchi yo'q bo'lsa, o'qituvchi tekshiruvini chetlab o'tib ishlayveradi.
    """
    start = grp_a.course.start_date
    end = grp_a.course.end_date

    teacher_free_slots = set()
    has_teacher = bool(grp_a.teacher_id)

    cur = start
    while cur <= end:
        if cur.weekday() <= 4:  # Dushanba - Juma
            if has_teacher:
                teacher_busy = set()
                for sc in GroupSchedule.objects.filter(date=cur, group__teacher=grp_a.teacher):
                    st = sc.start_time or sc.group.start_time
                    if st:
                        for i, (ps, _) in enumerate(PARA_TIMES):
                            if ps == st:
                                teacher_busy.add(i)
                for i in range(len(PARA_TIMES)):
                    if i not in teacher_busy:
                        teacher_free_slots.add((cur, i))
            else:
                # O'qituvchi bo'lmasa, barcha vaqtlar ochiq deb hisoblanadi
                for i in range(len(PARA_TIMES)):
                    teacher_free_slots.add((cur, i))
        cur += timedelta(days=1)

    if not teacher_free_slots:
        return None

    students_a = list(grp_a.students.all())
    student_a_ids = set(s.id for s in students_a)
    block_counts = defaultdict(int)

    for sc in GroupSchedule.objects.filter(
            date__range=(start, end),
            group__students__id__in=student_a_ids,
    ).prefetch_related('group__students'):
        st = sc.start_time or sc.group.start_time
        if st:
            for i, (ps, _) in enumerate(PARA_TIMES):
                if ps == st and (sc.date, i) in teacher_free_slots:
                    for s in sc.group.students.all():
                        if s.id in student_a_ids:
                            block_counts[s.id] += 1

    if not block_counts:
        return None

    # Eng ko'p ziddiyatga uchrayotgan talabani aniqlaymiz
    bad_id = max(block_counts, key=block_counts.get)
    bad_student = Student.objects.get(id=bad_id)

    # MUHIM: aniq matn tengligi (`language=grp_a.language`) o'rniga endi
    # TIL KESISHMASI tekshiriladi. Chunki guruh tili endi talabalar tarkibiga
    # qarab avtomatik hisoblanadi (`sync_group_language`) va aralash bo'lishi
    # mumkin (masalan 'uz-ru'). Aniq tenglik bo'lganda sof 'uz' guruh bilan
    # aralash 'uz-ru' guruh bir-biriga umuman mos kelmay qolardi — bu esa
    # asossiz ravishda ko'plab haqiqiy nomzodlarni chetlab o'tishga olib kelardi.
    grp_a_langs = parse_group_langs(grp_a.language)
    all_subject_groups = list(
        CourseGroup.objects.filter(
            course__subject=grp_a.course.subject,
        ).exclude(pk=grp_a.pk).prefetch_related('students')
    )
    lang_matched_groups = [
        g for g in all_subject_groups
        if parse_group_langs(g.language) & grp_a_langs
    ]

    # MUHIM: bad_student ning O'ZI ham yangi guruhga (p_grp) mos kelishi kerak —
    # avval faqat "candidate" (qaytaruvchi talaba) tekshirilar edi, bad_student
    # ning o'zi p_grp bilan to'qnashib qolishi mumkinligi HECH tekshirilmagan edi!
    bad_student_full_busy = set(
        GroupSchedule.objects.filter(group__students=bad_student)
        .exclude(group=grp_a)
        .values_list('date', 'start_time')
    )

    def _find_safe_swap(search_groups):
        """Berilgan guruhlar ro'yxatidan xavfsiz swap (safe_candidate + grp_b) qidiradi."""
        for p_grp in search_groups:
            p_grp_times = set(
                GroupSchedule.objects.filter(group=p_grp).values_list('date', 'start_time')
            )
            if bad_student_full_busy & p_grp_times:
                # bad_student ning o'zi shu guruhga to'g'ri kelmaydi — keyingisiga o'tamiz
                continue

            for candidate in p_grp.students.all():
                if candidate.id in student_a_ids:
                    continue

                cand_busy = set()
                for sc in GroupSchedule.objects.filter(
                        group__students=candidate,
                        date__range=(start, end)
                ):
                    st = sc.start_time or sc.group.start_time
                    if st:
                        for i, (ps, _) in enumerate(PARA_TIMES):
                            if ps == st:
                                cand_busy.add((sc.date, i))

                if not (teacher_free_slots & cand_busy):
                    return candidate, p_grp
        return None, None

    def _find_one_way(search_groups):
        """Berilgan guruhlar ro'yxatidan bir tomonlama ko'chirish uchun bo'sh
        VA hajmi MAX_GROUP_SIZE dan oshib ketmaydigan guruh qidiradi."""
        bad_busy_times = set(
            GroupSchedule.objects.filter(group__students=bad_student)
            .exclude(group=grp_a)
            .values_list('date', 'start_time')
        )
        for p_grp in search_groups:
            # YANGI TEKSHIRUV: bir tomonlama ko'chirish guruhni MAX_GROUP_SIZE
            # dan oshirib yubormasligi kerak (avval bu tekshirilmagan edi)
            if p_grp.students.count() + 1 > MAX_GROUP_SIZE:
                continue
            p_grp_times = set(
                GroupSchedule.objects.filter(group=p_grp)
                .values_list('date', 'start_time')
            )
            if not (bad_busy_times & p_grp_times):
                return p_grp
        return None

    # 1-URINISH: til mos keladigan guruhlar orasidan qidiramiz (afzal variant)
    safe_candidate, grp_b = _find_safe_swap(lang_matched_groups)
    used_cross_language = False

    # 2-OXIRGI IMKON: til mos kelmasa ham, HECH bo'lmaganda joy topilishi uchun
    # BARCHA fan guruhlaridan (tildan qat'iy nazar) qidiramiz. Bu faqat til mos
    # guruhlar orasida yechim topilmagandagina ishga tushadi.
    if not safe_candidate and len(lang_matched_groups) < len(all_subject_groups):
        cross_lang_groups = [g for g in all_subject_groups if g not in lang_matched_groups]
        safe_candidate, grp_b = _find_safe_swap(cross_lang_groups)
        if safe_candidate:
            used_cross_language = True

    # Agar o'rniga qaytadigan (safe_candidate) topilmasa, bir tomonlama ko'chirishga harakat qilamiz
    if not safe_candidate:
        fallback_grp_b = None
        # O'quvchilar MIN_GROUP_SIZE dan kam bo'lib qolmasligi uchun
        if (grp_a.students.count() - 1) >= MIN_GROUP_SIZE:
            # 1-URINISH: til mos keladigan guruhlar orasidan
            fallback_grp_b = _find_one_way(lang_matched_groups)
            # 2-OXIRGI IMKON: hech biri topilmasa, tildan qat'iy nazar barcha guruhlardan
            if not fallback_grp_b and len(lang_matched_groups) < len(all_subject_groups):
                cross_lang_groups = [g for g in all_subject_groups if g not in lang_matched_groups]
                fallback_grp_b = _find_one_way(cross_lang_groups)
                if fallback_grp_b:
                    used_cross_language = True

        if fallback_grp_b:
            with transaction.atomic():
                grp_a.students.remove(bad_student)
                fallback_grp_b.students.add(bad_student)
                sync_group_language(grp_a)
                sync_group_language(fallback_grp_b)

            teacher_status = "" if has_teacher else " (O'qituvchi belgilanmagan)"
            lang_note = " ⚠️ (til mos guruh topilmagani uchun boshqa tildagi guruhga ko'chirildi)" if used_cross_language else ""
            return (
                f"✅ Avtomatik ko'chirish{teacher_status}: '{grp_a.course.subject}' "
                f"{bad_student.first_name} → {fallback_grp_b.group_number}-guruhga ko'chirildi "
                f"(guruh tarkibi 9 tadan kam bo'lib qolmadi).{lang_note}"
            )
        return None

    # Muvaffaqiyatli almashtirish (Swap)
    with transaction.atomic():
        grp_a.students.remove(bad_student)
        grp_b.students.add(bad_student)
        grp_b.students.remove(safe_candidate)
        grp_a.students.add(safe_candidate)
        sync_group_language(grp_a)
        sync_group_language(grp_b)

    teacher_status = "" if has_teacher else " (O'qituvchi belgilanmagan)"
    lang_note = " ⚠️ (til mos guruh topilmagani uchun boshqa tildagi guruh bilan almashtirildi)" if used_cross_language else ""
    return (
        f"✅ Avtomatik almashtirish{teacher_status}: '{grp_a.course.subject}' "
        f"{bad_student.first_name} → {grp_b.group_number}-guruhga, "
        f"{safe_candidate.first_name} → {grp_a.group_number}-guruhga almashtirildi.{lang_note}"
    )


def split_subjects(raw):
    results = []
    current = ""
    depth = 0
    for char in str(raw):
        if char == '(':
            depth += 1
            current += char
        elif char == ')':
            depth -= 1
            current += char
        elif char == ';' and depth == 0:
            if current.strip():
                results.append(current.strip())
            current = ""
        else:
            current += char
    if current.strip():
        results.append(current.strip())
    return results


def get_weekly_schedule_data(week_start=None):
    today = dt_date.today()
    if week_start is None:
        week_start = today - timedelta(days=today.weekday())
    week_end = week_start + timedelta(days=5)

    groups = CourseGroup.objects.filter(
        is_scheduled=True
    ).select_related('course__subject', 'teacher').prefetch_related('schedule')

    grid_lists = defaultdict(list)

    for grp in groups:
        subject_name = str(grp.course.subject)
        # ── teacher None bo'lishi mumkin ──
        if grp.teacher:
            teacher_name = f"{grp.teacher.first_name} {grp.teacher.last_name}"
        else:
            teacher_name = "O'qituvchi biriktirilmagan"

        for sched in grp.schedule.filter(date__gte=week_start, date__lte=week_end):
            weekday = sched.date.weekday()
            if weekday > 5:
                continue
            st = sched.start_time or grp.start_time
            if not st:
                continue
            start_str = st.strftime("%H:%M")
            para_idx = next(
                (i for i, (s, e) in enumerate(PARA_TIMES_WEEKLY) if s == start_str), None
            )
            if para_idx is None:
                continue

            grid_lists[(weekday, para_idx)].append({
                'subject'     : subject_name,
                'teacher'     : teacher_name,
                'room'        : str(grp.room) if grp.room else '',
                'sched_id'    : sched.pk,
                'group_number': grp.group_number,
            })

    max_cols = max((len(v) for v in grid_lists.values()), default=0)

    grid = {}
    for (weekday, para_idx), items in grid_lists.items():
        for col, item in enumerate(items, 1):
            grid[(weekday, para_idx, col)] = item

    return {
        'max_group' : max_cols,
        'grid'      : grid,
        'week_start': week_start,
        'week_end'  : week_end,
    }


# ─────────────────────────────────────────
# LESSON LIST
# ─────────────────────────────────────────
@login_required
def lesson_list(request):
    if is_student(request.user):
        return redirect('student_dashboard')
    if is_teacher(request.user) and not is_admin(request.user):
        return redirect('teacher_dashboard')

    q = request.GET.get('q', '').strip()
    courses = Course.objects.select_related('subject').prefetch_related('groups').all()
    if q:
        courses = courses.filter(subject__name__icontains=q)

    courses_data = []
    for course in courses:
        total = course.groups.count()
        scheduled = course.groups.filter(is_scheduled=True).count()

        # Kurs guruhlari ichida o'qituvchisi biriktirilmagan guruh bormi?
        # Agar kamida 1 ta o'qituvchisiz guruh bo'lsa True, hammasida bo'lsa False qaytaradi
        has_unassigned = course.groups.filter(teacher__isnull=True).exists()

        courses_data.append({
            'course': course,
            'total_groups': total,
            'scheduled_groups': scheduled,
            'has_unassigned_teachers': has_unassigned,  # HTML shablonimiz uchun yangi flag
        })

    return render(request, "raspisaniya/lesson_list.html", {"courses_data": courses_data, "q": q})


@login_required
def lesson_create(request):
    # ── STEP 1 ──
    if request.method == "GET":
        all_subjects = Subject.objects.all()
        subjects_data = []
        for subj in all_subjects:
            count = Student.objects.filter(debts=subj).count()
            if count >= 10:
                subjects_data.append({'subject': subj, 'student_count': count})
        return render(request, "raspisaniya/lesson_create.html", {
            "step": 1,
            "subjects_data": subjects_data,
        })

    # ── STEP 2 ──
    if request.method == "POST" and request.POST.get("step") == "2":
        subject_id = request.POST.get("subject")
        subject = get_object_or_404(Subject, id=subject_id)

        start_date_raw = request.POST.get("start_date")
        total_lessons = request.POST.get("total_lessons")
        lessons_per_week = request.POST.get("lessons_per_week")
        include_saturday = request.POST.get("include_saturday", "0")

        if not all([start_date_raw, total_lessons, lessons_per_week]):
            messages.error(request, "Barcha maydonlarni to'ldiring")
            return redirect("lesson_create")

        total_lessons = int(total_lessons)
        lessons_per_week = int(lessons_per_week)
        start_date = parse_date(start_date_raw)

        # ── YANGI TEKSHIRUV: jadval tuzuvchi algoritm (find_schedule_for_group)
        # kurs turiga (24/16/8 paralik) qarab HAR DOIM qat'iy belgilangan
        # haftalik dars sonini qo'llaydi. Agar admin "Haftada necha marta"ga
        # boshqa son kiritgan bo'lsa, tugash sanasi noto'g'ri hisoblanib
        # qolardi — shuning uchun bu yerda avtomatik to'g'ri qiymatga
        # tuzatib, adminga ogohlantirish beramiz ──
        expected_per_week = get_expected_lessons_per_week(total_lessons)
        if lessons_per_week != expected_per_week:
            messages.warning(
                request,
                f"⚠ Diqqat: {total_lessons} paralik kurs uchun tizim haftada "
                f"faqat {expected_per_week} para joylashtira oladi (siz {lessons_per_week} "
                f"kiritgan edingiz). Tugash sanasi shunga qarab to'g'rilandi."
            )
            lessons_per_week = expected_per_week

        weeks_needed = math.ceil(total_lessons / lessons_per_week)
        end_date = start_date + timedelta(weeks=weeks_needed)
        end_date_raw = end_date.strftime("%Y-%m-%d")

        all_students = list(Student.objects.filter(debts=subject).distinct())
        if not all_students:
            messages.error(request, "Bu fandan yiqilgan o'quvchi yo'q")
            return redirect("lesson_create")

        # ── Tilga qarab ajratamiz ──
        students_by_lang = defaultdict(list)
        for st in all_students:
            students_by_lang[st.language].append(st)

        all_groups = []
        skipped_msgs = []
        group_index = 0

        for lang in sorted(students_by_lang.keys()):
            lang_students = students_by_lang[lang]
            lang_name = dict(LANGUAGE_CHOICES).get(lang, lang)
            groups = split_into_groups(lang_students)

            for g in groups:
                is_small = len(g) < 10
                if is_small:
                    skipped_msgs.append(
                        f"{lang_name} tili: {len(g)} ta o'quvchi "
                        f"(10 tadan kam, guruh shakillantirilmadi)"
                    )
                all_groups.append({
                    'index': group_index,
                    'lang': lang,
                    'lang_name': lang_name,
                    'students': g,
                    'is_small': is_small,
                })
                group_index += 1

        if not all_groups:
            messages.error(request, "Bu fandan o'quvchi yo'q")
            return redirect("lesson_create")

        groups_count = len(all_groups)

        # ── O'qituvchi bu yerda YO'Q ──
        return render(request, "raspisaniya/lesson_create.html", {
            "step": 2,
            "subject": subject,
            "all_groups": all_groups,
            "groups_count": groups_count,
            "start_date": start_date_raw,
            "end_date": end_date_raw,
            "total_lessons": total_lessons,
            "lessons_per_week": lessons_per_week,
            "skipped_langs": skipped_msgs,
            "all_students": all_students,
            "include_saturday": include_saturday,
        })

    # ── STEP 3 ──
    if request.method == "POST" and request.POST.get("step") == "3":
        subject_id = request.POST.get("subject_id")
        subject = get_object_or_404(Subject, id=subject_id)

        start_date_raw = request.POST.get("start_date")
        end_date_raw = request.POST.get("end_date")
        total_lessons = int(request.POST.get("total_lessons"))
        lessons_per_week = int(request.POST.get("lessons_per_week"))
        groups_count = int(request.POST.get("groups_count", 1))
        include_saturday = request.POST.get("include_saturday", "0") == "1"

        # ── Ehtiyot chorasi: STEP 2 da to'g'rilangan bo'lsa ham, formani
        # kimdir qo'lda o'zgartirib yubormasligi uchun bu yerda YANA bir bor
        # tekshiramiz ──
        lessons_per_week = get_expected_lessons_per_week(total_lessons)

        start_date = parse_date(start_date_raw)
        end_date = parse_date(end_date_raw)

        # ── Indekslarni students_ orqali aniqlaymiz
        #    (teacher_ endi yo'q) ──────────────────
        all_indices = []
        for key in request.POST.keys():
            if key.startswith("students_"):
                try:
                    all_indices.append(int(key.split("_", 1)[1]))
                except ValueError:
                    pass
        all_indices = sorted(set(all_indices))

        with transaction.atomic():
            course = Course.objects.create(
                subject=subject,
                start_date=start_date,
                end_date=end_date,
                total_lessons=total_lessons,
                lessons_per_week=lessons_per_week,
                lesson_duration=80,
                include_saturday=include_saturday,
            )

            group_number = 1
            for i in all_indices:
                selected_ids = request.POST.getlist(f"students_{i}")
                if not selected_ids:
                    continue
                selected_students = list(
                    Student.objects.filter(id__in=selected_ids)
                )
                if not selected_students:
                    continue

                lang = request.POST.get(
                    f"lang_{i}", selected_students[0].language
                )

                # ── O'qituvchisiz guruh — keyinroq biriktiriladi ──
                # MUHIM: 10 tadan kam bo'lsa ham guruh HAR DOIM saqlanadi
                # ("Ruxsat berish" talabi olib tashlandi — eski xatti-harakatga qaytarildi)
                cgroup = CourseGroup.objects.create(
                    course=course,
                    teacher=None,
                    group_number=group_number,
                    start_time=None,
                    weekdays=[],
                    language=lang,
                    is_scheduled=False,
                )
                cgroup.students.set(selected_students)
                sync_group_language(cgroup)

                for st in selected_students:
                    st.debts.remove(subject)

                group_number += 1

        messages.success(
            request,
            "Kurs yaratildi! Endi har bir guruhga o'qituvchi "
            "biriktiring, so'ng 'Jadval tuzish' tugmasini bosing."
        )
        return redirect("lesson_list")


# ─────────────────────────────────────────
# LESSON SCHEDULE
# ─────────────────────────────────────────
@login_required
def lesson_schedule(request, pk):
    course = get_object_or_404(Course, pk=pk)
    # Guruhlarni va ularning jadvallarini yuklab olamiz
    groups = course.groups.prefetch_related('students', 'schedule').select_related('teacher')
    duration = timedelta(minutes=80)

    # 1. Shu kursdagi barcha guruhlarda o'qiyotgan talabalar ID lari (takrorlanmaslik uchun)
    all_group_student_ids = set()
    for grp in groups:
        for s in grp.students.all():
            all_group_student_ids.add(s.id)

    # Shu fandan qarzdor va hali hech qaysi guruhga qo'shilmagan talabalar
    addable_students = Student.objects.filter(
        debts=course.subject
    ).exclude(
        id__in=all_group_student_ids
    )

    # 2. Shu fanga ixtisoslashgan barcha o'qituvchilar ro'yxati (Dropdown uchun asos)
    course_teachers = Teacher.objects.filter(subjects=course.subject).order_by('first_name')

    # Hafta kunlari nomlari lug'ati (agar loyihangizda yuqorida bo'lsa o'chirib qo'yishingiz mumkin)
    WEEKDAY_NAMES = {0: 'Dushanba', 1: 'Seshanba', 2: 'Chorshanba', 3: 'Payshanba', 4: 'Juma', 5: 'Shanba'}

    groups_data = []
    for grp in groups:
        # Guruhning dars vaqtlari ro'yxatini shakllantiramiz
        group_schedules = grp.schedule.exclude(start_time__isnull=True).values('date', 'start_time')

        # ── O'QITUVCHILARNING BANDLIGINI TEKSHIRISH (Joriy guruh uchun) ──
        teachers_with_status = []
        for teacher in course_teachers:
            is_busy = False

            # Agar o'qituvchi joriy guruhning o'z o'qituvchisi bo'lsa, uni band deb ko'rsatmaymiz
            if grp.teacher == teacher:
                is_busy = False
            else:
                # O'qituvchining dars vaqtlarini joriy guruhniki bilan solishtiramiz
                for sched in group_schedules:
                    clash_exists = GroupSchedule.objects.filter(
                        date=sched['date'],
                        start_time=sched['start_time'],
                        group__teacher=teacher
                    ).exists()

                    if clash_exists:
                        is_busy = True
                        break  # Bitta to'qnashuv topilsa, ushbu o'qituvchi uchun tekshiruvni to'xtatamiz

            teachers_with_status.append({
                'id': teacher.id,
                'full_name': f"{teacher.first_name} {teacher.last_name}",
                'is_busy': is_busy
            })

        # Jadvallarni formatlash
        schedule_list = []
        for s in grp.schedule.all().order_by('lesson_number'):
            st = s.start_time or grp.start_time
            if st:
                end_t = (datetime.datetime.combine(s.date, st) + duration).time()
                start_str = st.strftime("%H:%M")
                end_str = end_t.strftime("%H:%M")
            else:
                start_str = "—"
                end_str = "—"

            schedule_list.append({
                "sched": s,
                "weekday": WEEKDAY_NAMES.get(s.date.weekday(), "") if st else "",
                "start_time": start_str,
                "end_time": end_str,
            })

        # Har bir guruh uchun unga xos bo'lgan o'qituvchilar ro'yxatini (band/bo'sh statusi bilan) qo'shamiz
        groups_data.append({
            "group": grp,
            "schedule_list": schedule_list,
            "addable_students": addable_students,
            "teachers_list": teachers_with_status,  # <--- Yangi qo'shilgan qism
        })

    return render(request, "raspisaniya/lesson_schedule.html", {
        "course": course,
        "groups_data": groups_data,
        "rooms": Room.objects.all().order_by('name'),
    })


@login_required
def add_student_to_group(request, group_pk):
    group = get_object_or_404(CourseGroup, pk=group_pk)
    if request.method == "POST":
        student_id = request.POST.get("student_id")
        if student_id:
            student = get_object_or_404(Student, pk=student_id)

            # ── YANGI TEKSHIRUV: agar guruh allaqachon jadvallangan bo'lsa,
            # talabaning boshqa guruhlardagi dars vaqtlari bilan to'qnashib
            # qolmasligini tekshiramiz — aks holda talaba bir vaqtda ikki
            # joyga tushib qolishi mumkin edi ──
            conflict = get_student_group_conflict(student, group)
            if conflict:
                conflict_date, conflict_time = conflict
                messages.error(
                    request,
                    f"❌ {student} allaqachon {conflict_date.strftime('%d.%m.%Y')} kuni "
                    f"{conflict_time.strftime('%H:%M')} da boshqa darsga band — guruhga qo'shilmadi."
                )
                return redirect("lesson_schedule", pk=group.course.pk)

            group.students.add(student)
            sync_group_language(group)
            student.debts.remove(group.course.subject)
            messages.success(request, f"{student} guruhga qo'shildi.")
    return redirect("lesson_schedule", pk=group.course.pk)


@login_required
def lesson_schedule_excel(request, pk):
    course = get_object_or_404(Course, pk=pk)
    duration = timedelta(minutes=80)

    wb = Workbook()
    first = True
    for grp in course.groups.prefetch_related('students', 'schedule').select_related('teacher'):
        if first:
            ws = wb.active
            ws.title = f"{grp.group_number}-guruh"
            first = False
        else:
            ws = wb.create_sheet(title=f"{grp.group_number}-guruh")

        ws.append(["#", "Sana", "Hafta kuni", "Boshlanish", "Tugash", "O'qituvchi"])
        for s in grp.schedule.all():
            if grp.start_time:
                end_t = (datetime.combine(s.date, grp.start_time) + duration).time()
                ws.append([
                    s.lesson_number,
                    s.date.strftime("%d.%m.%Y"),
                    WEEKDAYS.get(s.date.weekday(), ""),
                    grp.start_time.strftime("%H:%M"),
                    end_t.strftime("%H:%M"),
                    str(grp.teacher),
                ])

        ws2 = wb.create_sheet(title=f"{grp.group_number}-guruh talabalar")
        ws2.append(["#", "O'quvchi", "O'qituvchi"])
        for idx, st in enumerate(grp.students.all(), 1):
            ws2.append([idx, str(st), str(grp.teacher)])

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = f'attachment; filename="jadval_{course.pk}.xlsx"'
    wb.save(response)
    return response


@login_required
def change_lesson_time(request, sched_pk):
    sched = get_object_or_404(GroupSchedule, pk=sched_pk)
    if request.method == "POST":
        new_time = request.POST.get("start_time")
        new_date = request.POST.get("date")
        # ── YANGI: "apply_to_future" checkbox — agar belgilangan bo'lsa va
        # faqat vaqt o'zgargan bo'lsa (sana o'sha-o'sha), shu haftadagi
        # o'zgarish KEYINGI barcha haftalarga ham qo'llaniladi ──
        apply_to_future = request.POST.get("apply_to_future") in ("1", "true", "on")

        new_date_val = parse_date(new_date) if new_date else sched.date
        if new_time:
            h, m = map(int, new_time.split(":"))
            new_time_val = dtime(h, m)
        else:
            new_time_val = sched.start_time

        updated, skipped_dates = apply_lesson_time_change(
            sched, new_date_val, new_time_val, apply_to_future=apply_to_future
        )

        if updated:
            if apply_to_future and updated > 1:
                messages.success(
                    request,
                    f"✅ {updated} ta hafta uchun dars vaqti {new_time} ga o'zgartirildi "
                    f"(bu va keyingi barcha haftalar)."
                )
            else:
                messages.success(request, f"{new_date_val} dars vaqti o'zgartirildi")
        if skipped_dates:
            dates_str = ", ".join(d.strftime('%d.%m.%Y') for d in skipped_dates[:5])
            extra = len(skipped_dates) - 5
            if extra > 0:
                dates_str += f" va yana {extra} ta"
            messages.warning(
                request,
                f"⚠ Quyidagi sanalarda o'qituvchi yoki talabalar band bo'lgani uchun "
                f"o'zgartirilmadi: {dates_str}"
            )
        if not updated and not skipped_dates:
            messages.error(request, "Dars topilmadi yoki o'zgartirish uchun hech narsa yo'q.")
    return redirect("lesson_schedule", pk=sched.group.course.pk)


@login_required
def change_teacher(request, group_pk):
    group = get_object_or_404(CourseGroup, pk=group_pk)
    if request.method != "POST":
        return redirect("lesson_schedule", pk=group.course.pk)

    teacher_id = request.POST.get("teacher_id", "").strip()
    if not teacher_id:
        messages.error(request, "O'qituvchi tanlanmadi")
        return redirect("lesson_schedule", pk=group.course.pk)

    teacher = get_object_or_404(Teacher, pk=teacher_id)

    # ── Vaqt to'qnashuvi tekshiruvi ──────────────────────────────
    group_schedules = group.schedule.values('date', 'start_time')

    conflicts = []
    for sched in group_schedules:
        if not sched['start_time']:
            continue
        clash = GroupSchedule.objects.filter(
            date=sched['date'],
            start_time=sched['start_time'],
            group__teacher=teacher,
        ).exclude(group=group).select_related(
            'group__course__subject'
        ).first()

        if clash:
            conflicts.append(
                f"{sched['date'].strftime('%d.%m.%Y')} "
                f"{sched['start_time'].strftime('%H:%M')} — "
                f"{clash.group.course.subject} "
                f"({clash.group.group_number}-guruh)"
            )

    if conflicts:
        shown = conflicts[:3]
        extra = len(conflicts) - 3
        msg   = "; ".join(shown)
        if extra > 0:
            msg += f" va yana {extra} ta"
        messages.error(
            request,
            f"❌ {teacher.first_name} o'qituvchisi band: {msg}. "
            f"Boshqa o'qituvchi tanlang."
        )
        return redirect("lesson_schedule", pk=group.course.pk)

    # ── To'qnashuv yo'q — saqlash ────────────────────────────────
    old_teacher = group.teacher
    group.teacher = teacher
    group.save(update_fields=['teacher'])

    messages.success(
        request,
        f"✅ {group.group_number}-guruh o'qituvchisi "
        f"{'↳ ' + old_teacher.first_name + ' → ' if old_teacher else ''}"
        f"{teacher.first_name} ga o'zgartirildi."
    )
    return redirect("lesson_schedule", pk=group.course.pk)


# ─────────────────────────────────────────
# ✅ O'ZGARISH 8: lesson_delete — o'chirilganda talabalar
#    qaytib subject.debts ga qo'shiladi
# ─────────────────────────────────────────
@login_required
def lesson_delete(request, pk):
    lesson = get_object_or_404(Course, pk=pk)
    if request.method == "POST":
        subject = lesson.subject
        # Kursga biriktirilgan barcha guruhlar va ulardagi talabalarni qaytarish
        for group in lesson.groups.prefetch_related('students').all():
            for student in group.students.all():
                student.debts.add(subject)
        lesson.delete()
        messages.success(request, "Dars o'chirildi va talabalar qayta ro'yxatga qaytarildi")
        return redirect("lesson_list")
    return render(request, "raspisaniya/lesson_delete.html", {"lesson": lesson, "course": lesson})


@staff_member_required
def admin_group_grades(request, group_pk):
    """Admin: guruh talabalarining baholari va davomati."""
    from raspisaniya.models import Grade, Attendance
    group = get_object_or_404(CourseGroup, pk=group_pk)
    students = group.students.all().order_by('last_name', 'first_name')
    schedules = list(group.schedule.all().order_by('date'))
    total_lessons = len(schedules)

    grade_map = {g.student_id: g for g in Grade.objects.filter(course_group=group)}
    all_att = Attendance.objects.filter(schedule__group=group).values('student_id', 'schedule_id', 'is_present')
    att_map = {(a['student_id'], a['schedule_id']): a['is_present'] for a in all_att}

    rows = []
    for st in students:
        cells = []
        came = missed = 0
        for sched in schedules:
            val = att_map.get((st.pk, sched.pk))
            if val is True:
                came += 1
                cells.append('present')
            elif val is False:
                missed += 1
                cells.append('absent')
            else:
                cells.append('none')
        missed_percent = round(missed / total_lessons * 100) if total_lessons > 0 else 0
        is_blocked = missed_percent > 25 and not group.teacher_can_edit
        grade = grade_map.get(st.pk)
        total_grade = None
        if grade:
            vals = [v for v in [grade.midterm, grade.current, grade.final] if v is not None]
            total_grade = round(sum(vals), 1) if vals else None
        rows.append({
            'student': st,
            'cells': cells,
            'came': came,
            'missed': missed,
            'missed_percent': missed_percent,
            'is_blocked': is_blocked,
            'grade': grade,
            'total': total_grade,
        })

    return render(request, "raspisaniya/admin_group_grades.html", {
        "group": group,
        "schedules": schedules,
        "rows": rows,
        "total_lessons": total_lessons,
    })


@login_required
def move_one_student(request):
    """Bitta talabani boshqa guruhga ko'chirish."""
    if request.method == "POST":
        student_pk   = request.POST.get('student_pk')
        from_group_pk = request.POST.get('from_group_pk')
        to_group_pk  = request.POST.get('to_group_pk')

        try:
            student    = Student.objects.get(pk=student_pk)
            from_group = CourseGroup.objects.get(pk=from_group_pk)
            to_group   = CourseGroup.objects.get(pk=to_group_pk)

            # ── YANGI TEKSHIRUV: talabaning to_group vaqtlari bilan
            # (from_group'dagi darslardan tashqari) to'qnashuvi bo'lmasin ──
            conflict = get_student_group_conflict(student, to_group, exclude_group=from_group)
            if conflict:
                conflict_date, conflict_time = conflict
                messages.error(
                    request,
                    f"❌ {student.first_name} {conflict_date.strftime('%d.%m.%Y')} kuni "
                    f"{conflict_time.strftime('%H:%M')} da band — {to_group.group_number}-guruhga ko'chirilmadi."
                )
                return redirect('build_schedule')

            from_group.students.remove(student)
            to_group.students.add(student)
            sync_group_language(from_group)
            sync_group_language(to_group)
            messages.success(request, f"{student.first_name} → {to_group.group_number}-guruhga ko'chirildi.")
        except Exception as e:
            messages.error(request, f"Xato: {e}")

    return redirect('build_schedule')


@login_required
@transaction.atomic
def add_group_and_redistribute(request, course_pk, language):
    """
    Yangi bo'sh guruh qo'shish funksiyasi.
    Eski guruhlar, ularning o'qituvchilari, talabalari va
    mavjud jadvallariga umuman ta'sir qilmaydi.
    """
    if request.method != "POST":
        return redirect('lesson_schedule', pk=course_pk)

    course = get_object_or_404(Course, pk=course_pk)

    # Mavjud guruhlar ichida eng katta guruh raqamini topamiz
    existing_groups = CourseGroup.objects.filter(course=course).order_by('-group_number')
    max_num = existing_groups.first().group_number if existing_groups.exists() else 0

    # Yangi bo'sh guruh yaratamiz
    new_group = CourseGroup.objects.create(
        course=course,
        teacher=None,  # O'qituvchi yo'q
        group_number=max_num + 1,  # Keyingi raqam
        language=language,  # Tanlangan til
        is_scheduled=False,  # Jadval hali tuzilmagan
        start_time=None,
        weekdays=[]
    )

    messages.success(
        request,
        f"✅ {max_num + 1}-guruh muvaffaqiyatli qo'shildi. "
        f"Endi talabalarni boshqa guruhlardan o'chirib, bu guruhga qo'shishingiz mumkin."
    )

    return redirect('lesson_schedule', pk=course.pk)


@login_required
def delete_course_group(request, group_pk):
    """Guruhni o'chirish — talabalar debts ga qaytadi."""
    group = get_object_or_404(CourseGroup, pk=group_pk)
    course_pk = group.course.pk
    if request.method == "POST":
        subject = group.course.subject
        for student in group.students.all():
            student.debts.add(subject)
        group.schedule.all().delete()
        group.delete()
        messages.success(request, "Guruh o'chirildi, talabalar qayta ro'yxatga qaytarildi.")
    return redirect("lesson_schedule", pk=course_pk)

@login_required
def remove_student_from_group(request, group_pk, student_pk):
    group = get_object_or_404(CourseGroup, pk=group_pk)
    student = get_object_or_404(Student, pk=student_pk)
    if request.method == "POST":
        group.students.remove(student)
        sync_group_language(group)
        student.debts.add(group.course.subject)
        messages.success(request, f"{student} guruhdan o'chirildi va qayta ro'yxatga qo'shildi")
    return redirect("lesson_schedule", pk=group.course.pk)


# ─────────────────────────────────────────
# TEACHER
# ─────────────────────────────────────────
@login_required
def teacher_list(request):
    q = request.GET.get('q', '').strip()
    sort = request.GET.get('sort', 'name_asc')

    # 'groups' o'rniga modelingizga mos bo'lgan 'coursegroup_set' ni yuklaymiz
    teachers = Teacher.objects.prefetch_related('subjects', 'coursegroup_set').all()

    if q:
        teachers = teachers.filter(first_name__icontains=q) | \
                   teachers.filter(last_name__icontains=q) | \
                   teachers.filter(teacher_id__icontains=q) | \
                   teachers.filter(subjects__name__icontains=q)

    teachers = teachers.distinct()

    # ID bo'yicha saralashda T- dan keyingi raqamni olib raqam sifatida tartiblash
    if sort in ('id_asc', 'id_desc'):
        def extract_num(t):
            m = re.search(r'\d+', t.teacher_id or '')
            return int(m.group()) if m else 0
        reverse = (sort == 'id_desc')
        teachers = sorted(teachers, key=extract_num, reverse=reverse)
    elif sort == 'name_asc':
        teachers = teachers.order_by('first_name')
    elif sort == 'name_desc':
        teachers = teachers.order_by('-first_name')
    elif sort == 'subj_asc':
        teachers = teachers.order_by('subjects__name', 'first_name')
    elif sort == 'subj_desc':
        teachers = teachers.order_by('-subjects__name', 'first_name')
    else:
        teachers = teachers.order_by('first_name')

    # HTML shablonda oson ishlatishimiz uchun har bir o'qituvchining dars yuklamasini yig'ib chiqamiz
    teachers_data = []
    for teacher in teachers:
        # Keshdagi ma'lumotdan foydalanish uchun .count() o'rniga len(...all()) ishlatamiz.
        # Bu bazaga qayta-qayta SQL so'rov yuborilishining oldini oladi (N+1 muammosi yechimi).
        group_count = len(teacher.coursegroup_set.all())

        teachers_data.append({
            'teacher': teacher,
            'group_count': group_count,
            'has_groups': group_count > 0
        })

    subjects = Subject.objects.all().order_by('name')

    return render(request, 'raspisaniya/teacher_list.html', {
        'teachers_data': teachers_data,
        'q': q,
        'subjects': subjects,
        'selected_subject': request.GET.get('subject', ''),
        'sort': sort,
    })


@login_required
def teacher_create(request):
    if request.method == 'POST':
        form = TeacherForm(request.POST)
        if form.is_valid():
            teacher_id = request.POST.get("teacher_id", "").strip()
            password = request.POST.get("password", "").strip()

            if not teacher_id:
                messages.error(request, "Teacher ID kiritilmagan")
                return render(request, 'raspisaniya/teacher_create.html', {
                    'form': form, 'subjects': Subject.objects.all(), 'selected_subjects': [],
                })

            if User.objects.filter(username=teacher_id).exists():
                messages.error(request, f"Bu ID ({teacher_id}) allaqachon mavjud")
                return render(request, 'raspisaniya/teacher_create.html', {
                    'form': form, 'subjects': Subject.objects.all(), 'selected_subjects': [],
                })

            with transaction.atomic():
                teacher = form.save(commit=False)
                teacher.teacher_id = teacher_id
                teacher.save()
                form.save_m2m()
                user = User.objects.create_user(
                    username=teacher_id,
                    password=password if password else teacher_id,
                    first_name=teacher.first_name,
                    last_name=teacher.last_name,
                )
                teacher.user = user
                teacher.save()

            messages.success(request, f"O'qituvchi qo'shildi. ID: {teacher_id}")
            return redirect('teacher_list')
    else:
        form = TeacherForm()
    return render(request, 'raspisaniya/teacher_create.html', {
        'form': form, 'subjects': Subject.objects.all(), 'selected_subjects': [],
    })


@login_required
def teacher_update(request, pk):
    teacher = get_object_or_404(Teacher, pk=pk)

    if request.method == 'POST':
        first_name = request.POST.get('first_name', '').strip()
        new_id = request.POST.get('teacher_id', '').strip()
        new_password = request.POST.get('new_password', '').strip()
        subject_ids = request.POST.getlist('subjects')

        # 1. Ismni tekshirish va yangilash
        if first_name:
            teacher.first_name = first_name
            if teacher.user:
                teacher.user.first_name = first_name
                teacher.user.save()

        # 2. XAVFSIZLIK TEKSHIRUVI: Yangi ID unikal ekanligini tekshirish
        if new_id and new_id != teacher.teacher_id:
            # Bazada aynan shu username mavjudligini tekshiramiz (o'zidan tashqari)
            if User.objects.filter(username=new_id).exists():
                messages.error(request,
                               f"Xatolik: Tizimda '{new_id}' ID ga ega foydalanuvchi allaqachon mavjud! Iltimos, boshqa ID kiriting.")

                # Xatolik bo'lgani uchun formani qayta yuklaymiz, foydalanuvchi kiritgan ma'lumotlar yo'qolmaydi
                return render(request, 'raspisaniya/teacher_update.html', {
                    'teacher': teacher,
                    'subjects': Subject.objects.all().order_by('name'),
                    'selected_subjects': [int(sid) for sid in subject_ids],  # Tanlangan fanlar o'chib ketmasligi uchun
                })

            # Agar muammo bo'lmasa, ID ni yangilaymiz
            teacher.teacher_id = new_id
            if teacher.user:
                teacher.user.username = new_id
                teacher.user.save()

        # 3. Parolni yangilash
        if new_password and teacher.user:
            if len(new_password) < 4:
                messages.error(request, "Xatolik: Yangi parol kamida 4 ta belgidan iborat bo'lishi kerak.")
                return render(request, 'raspisaniya/teacher_update.html', {
                    'teacher': teacher,
                    'subjects': Subject.objects.all().order_by('name'),
                    'selected_subjects': [int(sid) for sid in subject_ids],
                })
            teacher.user.set_password(new_password)
            teacher.user.save()

        # O'qituvchi modelini saqlash va fanlarni bog'lash
        teacher.save()
        teacher.subjects.set(subject_ids)

        messages.success(request, "O'qituvchi ma'lumotlari muvaffaqiyatli yangilandi.")
        return redirect('teacher_list')

    return render(request, 'raspisaniya/teacher_update.html', {
        'teacher': teacher,
        'subjects': Subject.objects.all().order_by('name'),
        'selected_subjects': list(teacher.subjects.values_list('id', flat=True)),
    })


@login_required
def teacher_delete(request, pk):
    teacher = get_object_or_404(Teacher, pk=pk)
    if request.method == 'POST':
        teacher.delete()
        messages.success(request, "O'qituvchi o'chirildi")
        return redirect('teacher_list')
    return render(request, 'raspisaniya/teacher_delete.html', {'teacher': teacher})


@login_required
def teacher_import(request):
    if request.method == "POST":
        form = TeacherImportForm(request.POST, request.FILES)
        if form.is_valid():
            file = request.FILES["file"]
            try:
                wb = load_workbook(file)
                ws = wb.active

                # 1. Barcha mavjud ma'lumotlarni bir marta xotiraga yuklaymiz
                # Fanlarni nomi bo'yicha lug'atga olamiz: {'Matematika': <Subject object>, ...}
                existing_subjects = {s.name: s for s in Subject.objects.all()}

                # Userlarning username'larini tezkor tekshirish uchun 'set'ga olamiz
                existing_usernames = set(User.objects.values_list('username', flat=True))

                with transaction.atomic():
                    for row in ws.iter_rows(min_row=2, values_only=True):
                        # Bo'sh qatorlarni tashlab ketish
                        if not row or not row[0] or not row[1]:
                            continue

                        tid = f"T-{str(row[0]).strip()}"
                        full_name = str(row[1]).strip()
                        if not full_name:
                            continue

                        first_name = full_name   # To'liq nom — qanday tursa shunday
                        last_name = ""

                        # 2. User yaratish yoki olish (Xotiradan tekshiramiz)
                        if tid not in existing_usernames:
                            u = User.objects.create_user(username=tid, password=tid)
                            existing_usernames.add(tid)
                        else:
                            u = User.objects.get(username=tid)

                        # 3. Teacher yaratish yoki yangilash
                        teacher, created = Teacher.objects.update_or_create(
                            user=u,
                            defaults={
                                'first_name': first_name,
                                'last_name': last_name,
                                'teacher_id': tid
                            }
                        )

                        # 4. Fanlarni bog'lash (Xotira orqali)
                        if len(row) > 2 and row[2]:
                            subject_names = [s.strip() for s in str(row[2]).split(",") if s.strip()]
                            for sname in subject_names:
                                # Agar fan lug'atda bo'lmasa, bazada yaratib lug'atga qo'shamiz
                                if sname not in existing_subjects:
                                    subj = Subject.objects.create(name=sname)
                                    existing_subjects[sname] = subj

                                # Fanlarni bog'laymiz
                                teacher.subjects.add(existing_subjects[sname])

                        teacher.save()

                messages.success(request, "O'qituvchilar muvaffaqiyatli import qilindi ✅")
                return redirect("teacher_list")
            except Exception as e:
                # Xatoni terminalda ham ko'rish uchun:
                print(f"Import Error: {e}")
                messages.error(request, f"Xatolik: {e}")
    else:
        form = TeacherImportForm()
    return render(request, "raspisaniya/teacher_import.html", {"form": form})


# ─────────────────────────────────────────
# STUDENT
# ─────────────────────────────────────────
@login_required
def student_list(request):
    q = request.GET.get('q', '').strip()

    # Saralash parametrlari
    sort_by = request.GET.get('sort_by', 'fio')  # fio, lang, subject
    direction = request.GET.get('direction', 'asc')  # asc yoki desc

    students = Student.objects.prefetch_related(
        'debts',
        'coursegroup_set__course__subject',
    ).select_related('group').all()

    # 1. Qidiruv filtri
    if q:
        students = students.filter(
            first_name__icontains=q
        ) | students.filter(
            last_name__icontains=q
        ) | students.filter(
            student_id__icontains=q
        ) | students.filter(
            group__name__icontains=q
        )
        students = students.distinct()

    # 2. Ma'lumotlarni yig'ish (Sizning eski tsiklingiz)
    students_data = []
    for student in students:
        completed = list({grp.course.subject for grp in student.coursegroup_set.all()})

        # Saralash oson bo'lishi uchun fanlar nomini bitta matnga birlashtiramiz (masalan: "Matematika, Fizika")
        subjects_text = ", ".join(sorted([subj.name for subj in completed]))

        students_data.append({
            'student': student,
            'completed': completed,
            'subjects_text': subjects_text  # Saralash uchun kerak
        })

    # 3. Python orqali mukammal saralash (Gibrid mantiq)
    is_reverse = (direction == 'desc')

    if sort_by == 'fio':
        # Familiya va Ism bo'yicha saralash
        students_data.sort(key=lambda x: (x['student'].last_name.lower(), x['student'].first_name.lower()),
                           reverse=is_reverse)

    elif sort_by == 'lang':
        # Ta'lim tili bo'yicha saralash (uz, ru, en...)
        students_data.sort(key=lambda x: (x['student'].language or '').lower(), reverse=is_reverse)

    elif sort_by == 'subject':
        # Biriktirilgan fanlar matni bo'yicha saralash
        students_data.sort(key=lambda x: x['subjects_text'].lower(), reverse=is_reverse)

    return render(request, 'raspisaniya/student_list.html', {
        'students_data': students_data,
        'q': q,
        'sort_by': sort_by,
        'direction': direction,
    })


@login_required
def student_create(request):
    if request.method == 'POST':
        form = StudentForm(request.POST)
        if form.is_valid():
            student_id = request.POST.get("student_id", "").strip()
            password = request.POST.get("password", "").strip()

            if not student_id:
                messages.error(request, "Student ID kiritilmagan")
                return render(request, 'raspisaniya/student_create.html', {
                    'form': form, 'subjects': Subject.objects.all(),
                    'groups': Group.objects.all(), 'selected_debts': [],
                })

            if User.objects.filter(username=student_id).exists():
                messages.error(request, f"Bu ID ({student_id}) allaqachon mavjud")
                return render(request, 'raspisaniya/student_create.html', {
                    'form': form, 'subjects': Subject.objects.all(),
                    'groups': Group.objects.all(), 'selected_debts': [],
                })

            with transaction.atomic():
                student = form.save(commit=False)
                student.student_id = student_id
                student.save()
                form.save_m2m()
                user = User.objects.create_user(
                    username=student_id,
                    password=password if password else student_id,
                    first_name=student.first_name,
                    last_name=student.last_name,
                )
                student.user = user
                student.save()

            messages.success(request, f"O'quvchi qo'shildi. ID: {student_id}")
            return redirect('student_list')
    else:
        form = StudentForm()
    return render(request, 'raspisaniya/student_create.html', {
        'form': form, 'subjects': Subject.objects.all(),
        'groups': Group.objects.all(), 'selected_debts': [],
    })


@login_required
def student_update(request, pk):
    student = get_object_or_404(Student, pk=pk)
    if request.method == 'POST':
        full_name = request.POST.get('first_name', '').strip()
        if full_name:
            student.first_name = full_name
            student.last_name = ''

        group_id = request.POST.get('group', '').strip()
        if group_id:
            try:
                student.group = Group.objects.get(pk=group_id)
            except Group.DoesNotExist:
                pass
        else:
            student.group = None

        language = request.POST.get('language', '').strip()
        if language:
            student.language = language

        student.save()

        debt_ids = request.POST.getlist('debts')
        student.debts.set(debt_ids)

        new_password = request.POST.get('new_password', '').strip()
        if new_password and student.user:
            student.user.set_password(new_password)
            student.user.save()
            messages.success(request, "O'quvchi va parol yangilandi")
        else:
            messages.success(request, "O'quvchi yangilandi")
        return redirect('student_list')

    form = StudentForm(instance=student)
    return render(request, 'raspisaniya/student_update.html', {
        'form': form,
        'student': student,
        'subjects': Subject.objects.all(),
        'groups': Group.objects.all(),
        'selected_debts': list(student.debts.values_list('id', flat=True)),
    })


@login_required
def admin_change_student_password(request, pk):
    student = get_object_or_404(Student, pk=pk)
    if request.method == "POST":
        new_password = request.POST.get("new_password", "").strip()
        if not new_password:
            messages.error(request, "Parol bo'sh bo'lmasin")
        elif not student.user:
            messages.error(request, "Bu talabaning tizim akkaunti yo'q")
        else:
            student.user.set_password(new_password)
            student.user.save()
            messages.success(request, f"{student} ning paroli o'zgartirildi")
    return redirect("student_list")


@login_required
def student_delete(request, pk):
    student = get_object_or_404(Student, pk=pk)
    if request.method == 'POST':
        student.delete()
        messages.success(request, "O'quvchi o'chirildi")
        return redirect('student_list')
    return render(request, 'raspisaniya/student_delete.html', {'student': student})


WORD_SUBJECTS_LOWER = [
    "noorganik kimyo",
    "organik kimyo",
    "fizik va kolloid kimyo",
    "analitik kimyo",
    "farmakognoziya",
    "farmatsevtik kimyo",
    "farmatsevtik texnologiya",
    "dorixonada ish yuritish",
    "sanoat texnologiyasi",
    "toksikologik kimyo",
    "sanoat farmatsiyasi",
    "farmatsevtik iqtisodiyoti",
]


def process_subject(raw_item):
    raw_item = raw_item.strip()
    if '(' in raw_item:
        name_only = raw_item[:raw_item.index('(')].strip()
    else:
        name_only = raw_item.strip()
    if name_only.lower() in WORD_SUBJECTS_LOWER:
        return raw_item
    else:
        return name_only


@login_required
@transaction.atomic
def import_students(request):
    if request.method == "POST":
        form = StudentImportForm(request.POST, request.FILES)
        if form.is_valid():
            file = request.FILES["file"]
            try:
                wb = load_workbook(file, read_only=True, data_only=True)
                ws = wb.active

                existing_users = {u.username: u for u in User.objects.filter(username__startswith='S-')}
                existing_groups = {g.name: g for g in Group.objects.all()}
                existing_subjects = {s.name: s for s in Subject.objects.all()}
                existing_students = {s.student_id: s for s in Student.objects.all()}

                new_users_to_create = []
                students_to_create = []
                students_to_update = []
                student_debts_collector = {}

                # Til mapping — Excel da qanday yozilsa shunga mos
                LANG_MAP = {
                    'o\'zbek': 'uz', 'uzbek': 'uz', 'uz': 'uz', "o'zbek": 'uz',
                    'рус': 'ru', 'rus': 'ru', 'ru': 'ru', 'russian': 'ru',
                    'қорақалпоқ': 'qq', 'karakalpak': 'qq', 'qq': 'qq',
                    'ingliz': 'en', 'english': 'en', 'en': 'en',
                }

                for row in ws.iter_rows(min_row=2, values_only=True):
                    if not row or not row[0] or not row[1]:
                        continue

                    sid = f"S-{str(row[0]).strip()}"
                    full_name = str(row[1]).strip()
                    if not full_name:
                        continue
                    first_name = full_name
                    last_name = ""

                    # Ta'lim tili — F ustuni (row[5])
                    lang_raw = str(row[5]).strip().lower() if len(row) > 5 and row[5] else 'uz'
                    language = LANG_MAP.get(lang_raw, 'uz')

                    # Guruh — E ustuni (row[4])
                    group_obj = None
                    if len(row) > 4 and row[4]:
                        g_name = str(row[4]).strip()
                        if g_name not in existing_groups:
                            group_obj = Group.objects.create(name=g_name)
                            existing_groups[g_name] = group_obj
                        else:
                            group_obj = existing_groups[g_name]

                    if sid not in existing_users:
                        user_obj = User(username=sid)
                        user_obj.set_password(sid)
                        new_users_to_create.append(user_obj)
                        existing_users[sid] = user_obj
                    else:
                        user_obj = existing_users[sid]

                    if sid not in existing_students:
                        st_obj = Student(
                            user=user_obj, student_id=sid,
                            first_name=first_name, last_name=last_name,
                            group=group_obj, language=language
                        )
                        students_to_create.append(st_obj)
                        existing_students[sid] = st_obj
                    else:
                        st_obj = existing_students[sid]
                        st_obj.first_name = first_name
                        st_obj.last_name = last_name
                        st_obj.group = group_obj
                        st_obj.language = language
                        if st_obj not in students_to_update:
                            students_to_update.append(st_obj)

                    # Fanlar — I ustuni (row[8])
                    if len(row) > 8 and row[8]:
                        raw_subjects = [s.strip() for s in str(row[8]).split(";") if s.strip()]
                        if sid not in student_debts_collector:
                            student_debts_collector[sid] = set()
                        for s_name in raw_subjects:
                            if s_name not in existing_subjects:
                                subj = Subject.objects.create(name=s_name)
                                existing_subjects[s_name] = subj
                            else:
                                subj = existing_subjects[s_name]
                            student_debts_collector[sid].add(subj)

                # Bulk create users
                if new_users_to_create:
                    User.objects.bulk_create(new_users_to_create, ignore_conflicts=True)
                    created_users = {u.username: u for u in
                                     User.objects.filter(username__in=[u.username for u in new_users_to_create])}
                    for s in students_to_create:
                        s.user = created_users.get(s.student_id)

                if students_to_create:
                    Student.objects.bulk_create(students_to_create)

                if students_to_update:
                    Student.objects.bulk_update(students_to_update, ['first_name', 'last_name', 'group', 'language'])

                if student_debts_collector:
                    db_students = {s.student_id: s for s in
                                   Student.objects.filter(student_id__in=student_debts_collector.keys())}
                    StudentDebtModel = Student.debts.through
                    debt_relations = []
                    for sid, subjects in student_debts_collector.items():
                        st_obj = db_students.get(sid)
                        if st_obj:
                            for subj_obj in subjects:
                                debt_relations.append(
                                    StudentDebtModel(student_id=st_obj.pk, subject_id=subj_obj.pk)
                                )
                    StudentDebtModel.objects.bulk_create(debt_relations, ignore_conflicts=True)

                messages.success(request, f"Import yakunlandi! {len(students_to_create)} yangi talaba qo'shildi.")
                return redirect("student_list")

            except Exception as e:
                print(f"IMPORT ERROR: {e}")
                messages.error(request, f"Xatolik yuz berdi: {str(e)}")
    else:
        form = StudentImportForm()

    return render(request, "raspisaniya/import_students.html", {"form": form})


# ─────────────────────────────────────────
# ROOM (XONA)
# ─────────────────────────────────────────
@login_required
def room_list(request):
    q = request.GET.get('q', '').strip()
    rooms = Room.objects.prefetch_related(
        'coursegroup_set__course__subject',
        'coursegroup_set__teacher',
    ).all().order_by('name')
    if q:
        rooms = rooms.filter(name__icontains=q)
    return render(request, 'raspisaniya/room_list.html', {'rooms': rooms, 'q': q})


@login_required
def room_create(request):
    if request.method == "POST":
        name = request.POST.get("name", "").strip()
        capacity = request.POST.get("capacity", 30)
        if not name:
            messages.error(request, "Xona nomi kiritilmagan")
        elif Room.objects.filter(name=name).exists():
            messages.error(request, f"'{name}' xonasi allaqachon mavjud")
        else:
            Room.objects.create(name=name, capacity=int(capacity))
            messages.success(request, f"'{name}' xonasi qo'shildi")
            return redirect("room_list")
    return render(request, 'raspisaniya/room_create.html')


@login_required
def room_delete(request, pk):
    room = get_object_or_404(Room, pk=pk)
    if request.method == "POST":
        room.delete()
        messages.success(request, "Xona o'chirildi")
    return redirect("room_list")


@login_required
def assign_room(request, group_pk):
    group = get_object_or_404(CourseGroup, pk=group_pk)
    if request.method == "POST":
        room_id = request.POST.get("room_id")
        if not room_id:
            group.room = None
            group.save()
            messages.success(request, "Xona biriktirilmadi (bo'shatildi)")
            return redirect("lesson_schedule", pk=group.course.pk)

        room = get_object_or_404(Room, pk=room_id)

        for sched in group.schedule.all():
            st = sched.start_time or group.start_time
            if not st:
                continue
            conflict = GroupSchedule.objects.filter(
                date=sched.date,
                start_time=st,
                group__room=room,
            ).exclude(group=group)
            if conflict.exists():
                conflict_grp = conflict.first().group
                messages.error(
                    request,
                    f"'{room.name}' xonasi {sched.date} kuni {st.strftime('%H:%M')} da "
                    f"'{conflict_grp.course.subject}' ({conflict_grp.group_number}-guruh) uchun band!"
                )
                return redirect("lesson_schedule", pk=group.course.pk)

        group.room = room
        group.save()
        messages.success(request, f"'{room.name}' xonasi biriktirildi")
    return redirect("lesson_schedule", pk=group.course.pk)


# ─────────────────────────────────────────
# SUBJECT
# ─────────────────────────────────────────
@login_required
def subject_list(request):
    q = request.GET.get('q', '').strip()

    # Saralash parametrlari (default holatda Fan nomi bo'yicha ASC)
    sort_by = request.GET.get('sort_by', 'subject')  # subject yoki debt
    direction = request.GET.get('direction', 'asc')  # asc yoki desc

    # Fanlarni va ularga tegishli qarzdorlarni bazadan yuklash
    subjects_queryset = Subject.objects.prefetch_related('debt_students').all()

    if q:
        subjects_queryset = subjects_queryset.filter(name__icontains=q)

    # Saralash mantiqi aniq ishlashi uchun ma'lumotlar ro'yxatini yig'amiz
    subjects_data = []
    for subj in subjects_queryset:
        subjects_data.append({
            'subject': subj,
            'debt_count': subj.debt_students.count()  # Qarzdorlar soni
        })

    # Python orqali mukammal saralash
    is_reverse = (direction == 'desc')

    if sort_by == 'subject':
        # Fan nomi bo'yicha saralash
        subjects_data.sort(key=lambda x: (x['subject'].name or '').lower(), reverse=is_reverse)

    elif sort_by == 'debt':
        # Qarzdor talabalar soni bo'yicha saralash
        subjects_data.sort(key=lambda x: x['debt_count'], reverse=is_reverse)

    return render(request, 'raspisaniya/subject_list.html', {
        'subjects_data': subjects_data,
        'q': q,
        'sort_by': sort_by,
        'direction': direction,
    })


@login_required
def subject_create(request):
    if request.method == 'POST':
        form = SubjectForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, "Fan qo'shildi")
            return redirect('subject_list')
    else:
        form = SubjectForm()
    return render(request, 'raspisaniya/subject_create.html', {'form': form})


@login_required
def subject_update(request, pk):
    subject = get_object_or_404(Subject, pk=pk)
    if request.method == 'POST':
        form = SubjectForm(request.POST, instance=subject)
        if form.is_valid():
            form.save()
            messages.success(request, "Fan yangilandi")
            return redirect('subject_list')
    else:
        form = SubjectForm(instance=subject)
    return render(request, 'raspisaniya/subject_create.html', {'form': form})


@login_required
def subject_delete(request, pk):
    subject = get_object_or_404(Subject, pk=pk)
    if request.method == 'POST':
        subject.delete()
        messages.success(request, "Fan o'chirildi")
        return redirect('subject_list')
    return render(request, 'raspisaniya/subject_delete.html', {'subject': subject})


@login_required
def subject_students(request, pk):
    subject = get_object_or_404(Subject, pk=pk)
    students = Student.objects.filter(debts=subject).order_by('last_name')
    return render(request, 'raspisaniya/subject_students.html', {
        'subject': subject, 'students': students,
    })


@login_required
def subject_students_excel(request, pk):
    subject = get_object_or_404(Subject, pk=pk)
    students = Student.objects.filter(debts=subject).order_by('last_name')

    wb = Workbook()
    ws = wb.active
    ws.title = subject.name
    ws.append(["#", "Familiya", "Ism Sharif", "Guruh"])
    for i, student in enumerate(students, 1):
        ws.append([i, student.last_name, student.first_name,
                   str(student.group) if student.group else "—"])

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = f'attachment; filename="{subject.name}_qarzdorlar.xlsx"'
    wb.save(response)
    return response


@login_required
def build_schedule(request):
    """
    To'liq, avtomatlashtirilgan dars jadvalini tuzish funksiyasi.
    while tsikli yordamida zanjirli optimallashtirish (bitta bosishda maksimal natija).
    """
    auto_resolve_messages = []
    group_last_positions = {}
    already_moved_students = set()
    total_success_count = 0

    def sort_key(g):
        total_lessons = g.course.total_lessons if g.course else 0
        student_count = len(g.students.all())
        return (-total_lessons, -student_count)

    def _rebuild_schedule(grp, course, teacher):
        same_subject_busy = list(
            GroupSchedule.objects.filter(
                group__course__subject=course.subject,
                group__is_scheduled=True,
            ).exclude(group=grp).values_list('date', 'start_time')
        )

        sched = find_schedule_for_group(
            course.start_date, course.end_date,
            course.total_lessons, course.lessons_per_week,
            teacher, list(grp.students.all()),
            group_number=grp.group_number,
            include_saturday=getattr(course, 'include_saturday', False),
            same_subject_busy=same_subject_busy,
        )
        fc = len(sched)
        cf = getattr(find_schedule_for_group, '_last_conflict_info', [])
        return sched, fc, cf

    # ── 🔄 MULTI-PASS OPTIMIZATION (WHILE TSIKLI) ──
    iteration = 0
    max_iterations = 10  # Cheksiz aylanib ketmasligi uchun xavfsizlik limiti

    while iteration < max_iterations:
        iteration += 1

        # Har bir yangi urinishda jadval qilinmagan guruhlarni qayta yuklaymiz
        unscheduled_groups = list(
            CourseGroup.objects.filter(
                is_scheduled=False
            ).select_related('course', 'course__subject', 'teacher')
            .prefetch_related('students')
        )

        if not unscheduled_groups:
            break

        unscheduled_list = sorted(unscheduled_groups, key=sort_key)
        iteration_success_count = 0  # Ushbu aylanishda nechta guruh muvaffaqiyatli joylashdi?

        for grp in unscheduled_list:
            if not CourseGroup.objects.filter(pk=grp.pk).exists():
                continue

            course = grp.course
            teacher = grp.teacher if grp.teacher_id else None

            schedule = []
            found_count = 0
            conflicts = []

            # ── 1-BOSQICH: To'g'ridan-to'g'ri joylashtirish ──
            find_schedule_for_group._last_conflict_info = []
            find_schedule_for_group._last_missing = 0
            find_schedule_for_group._last_no_slot_in_week = False

            schedule, found_count, conflicts = _rebuild_schedule(grp, course, teacher)

            # ── 2-BOSQICH: Joy topilmasa — 3 ta asosiy aqlli algoritmni ishga tushirish ──
            if found_count < course.total_lessons:
                MAX_ROUNDS = 15
                for _ in range(MAX_ROUNDS):
                    if found_count >= course.total_lessons:
                        break
                    changed = False

                    # 1. ALGORITM: Talaba almashtirish (Subject swap)
                    sid = transaction.savepoint()
                    resolved = _auto_resolve_conflicts_by_subject_swap(
                        grp, conflicts, already_moved_student_ids=already_moved_students
                    )
                    if resolved:
                        test_schedule, test_fc, test_cf = _rebuild_schedule(grp, course, teacher)
                        if test_fc <= found_count:
                            transaction.savepoint_rollback(sid)
                        else:
                            transaction.savepoint_commit(sid)
                            auto_resolve_messages.extend(resolved)
                            schedule, found_count, conflicts = test_schedule, test_fc, test_cf
                            changed = True
                            if found_count >= course.total_lessons:
                                break

                    # 2. ALGORITM: Parallel swap
                    sid = transaction.savepoint()
                    swap_msg = _auto_resolve_via_parallel_swap(grp)
                    if swap_msg:
                        test_schedule, test_fc, test_cf = _rebuild_schedule(grp, course, teacher)
                        if test_fc <= found_count:
                            transaction.savepoint_rollback(sid)
                        else:
                            transaction.savepoint_commit(sid)
                            auto_resolve_messages.append(swap_msg)
                            schedule, found_count, conflicts = test_schedule, test_fc, test_cf
                            changed = True
                            if found_count >= course.total_lessons:
                                break

                    # 3. ALGORITM: Boshqa fan guruhi bilan vaqt almashish
                    sid = transaction.savepoint()
                    cross_msg = _auto_resolve_via_cross_subject_swap(
                        grp, conflicts, group_last_positions=group_last_positions
                    )
                    if cross_msg:
                        test_schedule, test_fc, test_cf = _rebuild_schedule(grp, course, teacher)
                        if test_fc <= found_count:
                            transaction.savepoint_rollback(sid)
                        else:
                            transaction.savepoint_commit(sid)
                            auto_resolve_messages.append(cross_msg)
                            schedule, found_count, conflicts = test_schedule, test_fc, test_cf
                            changed = True
                            if found_count >= course.total_lessons:
                                break

                    if not changed:
                        break

                # ── 3-BOSQICH: Brute Force chorasi ──
                if found_count < course.total_lessons:
                    MAX_BRUTE_ROUNDS = 15
                    brute_used = 0

                    while found_count < course.total_lessons and brute_used < MAX_BRUTE_ROUNDS:
                        sid = transaction.savepoint()
                        brute_msg = _brute_force_find_slot(grp)

                        if brute_msg is None:
                            break

                        test_schedule, test_fc, test_cf = _rebuild_schedule(grp, course, teacher)
                        if test_fc <= found_count:
                            transaction.savepoint_rollback(sid)
                            break
                        else:
                            transaction.savepoint_commit(sid)
                            auto_resolve_messages.append(brute_msg)
                            schedule, found_count, conflicts = test_schedule, test_fc, test_cf
                            brute_used += 1

            # ── 💡 4-BOSQICH: MUAMMOLI TALABANI MAJBURIY SURISH ──
            if found_count < course.total_lessons:
                same_course_groups = list(course.groups.all())
                sid = transaction.savepoint()

                eviction_msg = _auto_resolve_by_force_student_eviction(grp, conflicts, same_course_groups)
                if eviction_msg:
                    test_schedule, test_fc, test_cf = _rebuild_schedule(grp, course, teacher)
                    if test_fc >= course.total_lessons:
                        transaction.savepoint_commit(sid)
                        auto_resolve_messages.append(eviction_msg)
                        schedule, found_count, conflicts = test_schedule, test_fc, test_cf
                    else:
                        transaction.savepoint_rollback(sid)

            # ── 5-BOSQICH: JADVAL SHAKLLANMASA GURUHNI DINAMIK TARQATAMIZ (DISSOLUTION) ──
            # Eslatma: while tsiklida birdan tarqatib yubormaslik uchun buni faqat oxirgi urinishda bajarish ham mumkin,
            # lekin mantiqiy zanjir buzilmasligi uchun hozircha o'z joyida qoldi.
            if found_count < course.total_lessons:
                same_course_groups = list(course.groups.all())
                dissolved, dissolve_msg = _try_dissolve_and_distribute_group(grp, same_course_groups)
                if dissolved:
                    auto_resolve_messages.append(dissolve_msg)
                    iteration_success_count += 1  # Tarqatish ham muvaffaqiyatli qadam
                    continue

            # ── JADVALNI SAQLASH BLOKI ──
            if found_count >= course.total_lessons:
                from collections import Counter
                para_counter = Counter(p_start for _, p_start, _ in schedule)
                most_common_para = para_counter.most_common(1)[0][0]
                grp.start_time = most_common_para
                grp.weekdays = list({d.weekday() for d, _, _ in schedule})
                grp.is_scheduled = True

                for attempt in range(5):
                    try:
                        with transaction.atomic():
                            grp.save()
                            GroupSchedule.objects.bulk_create([
                                GroupSchedule(
                                    group=grp, date=ld,
                                    lesson_number=idx, start_time=p_start
                                )
                                for idx, (ld, p_start, p_end) in enumerate(schedule, 1)
                                if not GroupSchedule.objects.filter(group=grp, date=ld, lesson_number=idx).exists()
                            ])
                        break
                    except Exception:
                        time.sleep(0.5)
                        continue

                # ── YANGI OGOHLANTIRISH: agar band kunlar ko'pligi sababli
                # oxirgi dars kursning e'lon qilingan tugash sanasidan (course.end_date)
                # keyin joylashgan bo'lsa — bu haqda ADMINGA OCHIQ xabar beramiz.
                # Aks holda kurs "17.08.2026da tugaydi" deb ko'rsatilgan bo'lsa-da,
                # aslida u sezilmasdan keyingi oylargacha davom etib ketishi mumkin edi.
                last_lesson_date = schedule[-1][0] if schedule else None
                if last_lesson_date and last_lesson_date > course.end_date:
                    overrun_days = (last_lesson_date - course.end_date).days
                    messages.warning(
                        request,
                        f"⏰ '{course.subject}' {grp.group_number}-guruh: band kunlar ko'pligi "
                        f"sababli oxirgi dars belgilangan tugash sanasidan "
                        f"({course.end_date.strftime('%d.%m.%Y')}) {overrun_days} kun keyin — "
                        f"{last_lesson_date.strftime('%d.%m.%Y')} da joylashdi. Kurs muddatini "
                        f"yoki band kunlarni ko'rib chiqishni tavsiya etamiz."
                    )

                iteration_success_count += 1
                total_success_count += 1

        # ── ⚠️ TSICLDAN CHIQISH SHARTI: ──
        # Agar bu aylanishda biror dona ham guruh jadvalga o'tira olmagan bo'lsa,
        # demak qo'shimcha urinishlardan foyda yo'q — cheksiz aylanishni oldini olish uchun chiqamiz.
        if iteration_success_count == 0:
            break

    # ── 🚨 XATOLIKLARNI YIG'ISH (FAQAT OXIRGI NUQTADA SIG'MAY QOLGANLAR UCHUN) ──
    final_unscheduled_groups = list(
        CourseGroup.objects.filter(is_scheduled=False)
        .select_related('course', 'course__subject', 'teacher')
        .prefetch_related('students')
    )

    if final_unscheduled_groups:
        error_details = []
        for grp in final_unscheduled_groups:
            course = grp.course
            teacher = grp.teacher if grp.teacher_id else None

            # Oxirgi holatdagi konfliktlarni tekshirish uchun qayta rebuild qilib olamiz
            schedule, found_count, conflicts = _rebuild_schedule(grp, course, teacher)
            no_slot = getattr(find_schedule_for_group, '_last_no_slot_in_week', False)

            other_groups = CourseGroup.objects.filter(
                course=course, is_scheduled=True,
            ).exclude(pk=grp.pk).prefetch_related('students')

            teacher_conflicts_display = []
            seen_teacher_groups = set()
            for c in conflicts:
                if c['type'] != 'teacher':
                    continue
                if c['group'].pk in seen_teacher_groups:
                    continue
                seen_teacher_groups.add(c['group'].pk)
                teacher_conflicts_display.append({
                    'date': c['date'],
                    'start_time': c['para_time'][0],
                    'subject': c['subject'],
                    'group': c['group'],
                })
                if len(teacher_conflicts_display) >= 10:
                    break

            student_groups = defaultdict(list)
            for c in conflicts:
                if c['type'] != 'student':
                    continue
                key = (c['date'], c['para_time'][0], c['group'].pk)
                student_groups[key].append(c)

            student_conflicts_display = []
            for key, items in list(student_groups.items())[:10]:
                first = items[0]
                all_students = []
                for it in items:
                    for st in it['busy_students']:
                        if st not in all_students:
                            all_students.append(st)
                student_conflicts_display.append({
                    'date': first['date'],
                    'start_time': first['para_time'][0],
                    'subject': first['subject'],
                    'group': first['group'],
                    'busy_students': all_students,
                })

            teacher_suggestion = None
            student_move_suggestions = []
            swap_suggestions = []

            # (Sizning mavjud suggestion mantiqlaringiz...)
            # ... [Bu yerda o'zgarishsiz qoladi] ...

            error_details.append({
                'group': grp,
                'course': course,
                'other_groups': other_groups,
                'found_count': found_count,
                'missing_count': course.total_lessons - found_count,
                'teacher_conflicts_display': teacher_conflicts_display,
                'student_conflicts_display': student_conflicts_display,
                'teacher_suggestion': teacher_suggestion,
                'student_move_suggestions': student_move_suggestions,
                'swap_suggestions': swap_suggestions,
                'no_teacher': not grp.teacher_id,
                'no_slot': no_slot,
            })

        if auto_resolve_messages:
            for msg in auto_resolve_messages:
                messages.info(request, msg)

        return render(request, "raspisaniya/build_schedule_errors.html", {
            "error_details": error_details,
            "success_count": total_success_count,
        })

    if auto_resolve_messages:
        for msg in auto_resolve_messages:
            messages.info(request, msg)

    messages.success(request, f"Jadval muvaffaqiyatli tuzildi! Jami {total_success_count} ta guruh.")
    return redirect("lesson_list")


def apply_teacher_suggestion(request, group_pk, teacher_pk):
    """Taklif: guruhga boshqa o'qituvchini biriktirish."""
    if request.method == "POST":
        grp = get_object_or_404(CourseGroup, pk=group_pk)
        teacher = get_object_or_404(Teacher, pk=teacher_pk)

        # ── YANGI TEKSHIRUV: agar guruh allaqachon jadvallangan bo'lsa
        # (masalan mavjud guruhga o'qituvchi almashtirilayotgan bo'lsa),
        # yangi o'qituvchining boshqa guruhlardagi vaqtlari bilan
        # to'qnashmasligini tekshiramiz ──
        conflict = get_teacher_group_conflict(teacher, grp)
        if conflict:
            conflict_date, conflict_time = conflict
            messages.error(
                request,
                f"❌ {teacher} {conflict_date.strftime('%d.%m.%Y')} kuni "
                f"{conflict_time.strftime('%H:%M')} da band — biriktirilmadi."
            )
            return redirect('build_schedule')

        grp.teacher = teacher
        grp.save()
        messages.success(
            request,
            f"'{grp.course.subject}' {grp.group_number}-guruh uchun o'qituvchi "
            f"{teacher} ga almashtirildi. Jadval qayta tuzilmoqda..."
        )
    return redirect('build_schedule')


@login_required
def apply_swap_suggestion(request):
    """Taklif: boshqa guruhning bir kunlik darsini boshqa vaqtga ko'chirish."""
    if request.method == "POST":
        group_pk = request.POST.get("group_pk")
        date_str = request.POST.get("date")
        old_time_str = request.POST.get("old_time")
        new_time_str = request.POST.get("new_time")

        grp = get_object_or_404(CourseGroup, pk=group_pk)
        d = parse_date(date_str)
        oh, om = map(int, old_time_str.split(":"))
        nh, nm = map(int, new_time_str.split(":"))
        old_t = dtime(oh, om)
        new_t = dtime(nh, nm)

        sched = GroupSchedule.objects.filter(group=grp, date=d, start_time=old_t).first()
        if sched:
            # ── YANGI TEKSHIRUV: yangi vaqtda (d, new_t) o'qituvchi yoki
            # talabalar boshqa darsga band emasligini tekshiramiz ──
            teacher_id = grp.teacher_id
            student_ids = list(grp.students.values_list('id', flat=True))

            if teacher_id and GroupSchedule.objects.filter(
                date=d, start_time=new_t, group__teacher_id=teacher_id,
            ).exclude(pk=sched.pk).exists():
                messages.error(
                    request,
                    f"❌ O'qituvchi {grp.teacher} {d.strftime('%d.%m.%Y')} kuni {new_time_str} da band — ko'chirilmadi."
                )
                return redirect('build_schedule')

            if student_ids and GroupSchedule.objects.filter(
                date=d, start_time=new_t, group__students__id__in=student_ids,
            ).exclude(pk=sched.pk).exists():
                messages.error(
                    request,
                    f"❌ Ba'zi talabalar {d.strftime('%d.%m.%Y')} kuni {new_time_str} da band — ko'chirilmadi."
                )
                return redirect('build_schedule')

            sched.start_time = new_t
            sched.save(update_fields=['start_time'])
            messages.success(
                request,
                f"'{grp.course.subject}' {grp.group_number}-guruhning {d.strftime('%d.%m.%Y')} "
                f"kungi darsi {old_time_str} dan {new_time_str} ga ko'chirildi. "
                f"Jadval qayta tuzilmoqda..."
            )
        else:
            messages.error(request, "Dars topilmadi — balki allaqachon o'zgargan. Qayta urinib ko'ring.")
    return redirect('build_schedule')


@login_required
def apply_student_swap_suggestion(request, group_pk):
    """
    Guruh jadvalini to'sib qo'ygan muammoli talabani aniqlash va
    parallel guruhdan zararsiz talabaga AVTOMATIK almashtirish.
    Tasdiqlash sahifasi yo'q — darhol bajariladi.
    """
    grp_a = get_object_or_404(CourseGroup, pk=group_pk)
    start = grp_a.course.start_date
    end   = grp_a.course.end_date

    # 1. O'qituvchining bo'sh slotlari
    teacher_free_slots = set()
    cur = start
    while cur <= end:
        if cur.weekday() <= 4:
            teacher_busy = set()
            for sc in GroupSchedule.objects.filter(date=cur, group__teacher=grp_a.teacher):
                st = sc.start_time or sc.group.start_time
                if st:
                    for i, (ps, _) in enumerate(PARA_TIMES):
                        if ps == st:
                            teacher_busy.add(i)
            for i in range(len(PARA_TIMES)):
                if i not in teacher_busy:
                    teacher_free_slots.add((cur, i))
        cur += timedelta(days=1)

    if not teacher_free_slots:
        messages.error(request,
            f"O'qituvchi {grp_a.teacher} da umuman bo'sh vaqt yo'q!")
        return redirect('build_schedule')

    # 2. Har bir talabaning nechta slotni to'sayotganini hisoblash
    students_a    = list(grp_a.students.all())
    student_a_ids = set(s.id for s in students_a)
    block_counts  = defaultdict(int)

    for sc in GroupSchedule.objects.filter(
        date__range=(start, end),
        group__students__id__in=student_a_ids,
    ).prefetch_related('group__students'):
        st = sc.start_time or sc.group.start_time
        if st:
            for i, (ps, _) in enumerate(PARA_TIMES):
                if ps == st and (sc.date, i) in teacher_free_slots:
                    for s in sc.group.students.all():
                        if s.id in student_a_ids:
                            block_counts[s.id] += 1

    if not block_counts:
        messages.warning(request,
            "Bu guruh talabalarida konflikt aniqlanmadi.")
        return redirect('build_schedule')

    # Eng ko'p to'sayotgan talaba
    bad_id      = max(block_counts, key=block_counts.get)
    bad_student = Student.objects.get(id=bad_id)

    # 3. Parallel guruhlardan zararsiz talaba qidirish
    parallel_groups = CourseGroup.objects.filter(
        course__subject=grp_a.course.subject
    ).exclude(pk=grp_a.pk).prefetch_related('students')

    safe_candidate = None
    grp_b          = None

    # MUHIM: bad_student ning O'ZI ham yangi guruhga (p_grp) mos kelishi kerak
    bad_student_full_busy = set(
        GroupSchedule.objects.filter(group__students=bad_student)
        .exclude(group=grp_a)
        .values_list('date', 'start_time')
    )

    for p_grp in parallel_groups:
        p_grp_times = set(
            GroupSchedule.objects.filter(group=p_grp).values_list('date', 'start_time')
        )
        if bad_student_full_busy & p_grp_times:
            continue

        for candidate in p_grp.students.all():
            if candidate.id in student_a_ids:
                continue

            # Nomzodning band slotlari
            cand_busy = set()
            for sc in GroupSchedule.objects.filter(
                group__students=candidate,
                date__range=(start, end)
            ):
                st = sc.start_time or sc.group.start_time
                if st:
                    for i, (ps, _) in enumerate(PARA_TIMES):
                        if ps == st:
                            cand_busy.add((sc.date, i))

            # O'qituvchi bo'sh slotlarining HECH BIRIDA band bo'lmasligi kerak
            if not (teacher_free_slots & cand_busy):
                safe_candidate = candidate
                grp_b          = p_grp
                break
        if safe_candidate:
            break

    # 4. Avtomatik almashtirish — tasdiqlashsiz
    if safe_candidate and grp_b:
        with transaction.atomic():
            grp_a.students.remove(bad_student)
            grp_b.students.add(bad_student)
            grp_b.students.remove(safe_candidate)
            grp_a.students.add(safe_candidate)
            sync_group_language(grp_a)
            sync_group_language(grp_b)

        messages.success(request,
            f"✅ Avtomatik almashtirish bajarildi: "
            f"{bad_student.first_name} → {grp_b.group_number}-guruhga, "
            f"{safe_candidate.first_name} → {grp_a.group_number}-guruhga ko'chirildi. "
            f"Jadval qayta tuzilmoqda..."
        )
    else:
        messages.error(request,
            f"Parallel guruhlarda o'qituvchi {grp_a.teacher} vaqtiga "
            f"mos keladigan zararsiz talaba topilmadi. "
            f"Muddatni uzaytirishni yoki boshqa o'qituvchi tanlashni tavsiya etamiz."
        )

    return redirect('build_schedule')



@login_required
def move_students(request, from_group_pk, to_group_pk):
    from_group = get_object_or_404(CourseGroup, pk=from_group_pk)
    to_group = get_object_or_404(CourseGroup, pk=to_group_pk)

    if request.method == "POST":
        student_ids = request.POST.getlist("student_ids")
        students = list(from_group.students.filter(id__in=student_ids))

        # ── YANGI TEKSHIRUV: har bir talaba uchun to_group vaqtlari bilan
        # to'qnashuvni alohida tekshiramiz — to'qnashgan talabalar
        # ko'chirilmaydi, faqat toza (band bo'lmagan) talabalar ko'chadi ──
        moved_students = []
        skipped_students = []
        for st in students:
            conflict = get_student_group_conflict(st, to_group, exclude_group=from_group)
            if conflict:
                skipped_students.append((st, conflict))
            else:
                from_group.students.remove(st)
                to_group.students.add(st)
                moved_students.append(st)

        sync_group_language(from_group)
        sync_group_language(to_group)

        if moved_students:
            messages.success(request, f"{len(moved_students)} ta talaba ko'chirildi.")
        if skipped_students:
            names = ", ".join(
                f"{st.first_name} ({d.strftime('%d.%m.%Y')} {t.strftime('%H:%M')} da band)"
                for st, (d, t) in skipped_students[:5]
            )
            extra = len(skipped_students) - 5
            if extra > 0:
                names += f" va yana {extra} ta"
            messages.error(request, f"❌ Quyidagi talabalar vaqt to'qnashuvi sababli ko'chirilmadi: {names}")
        return redirect("build_schedule")

    return render(request, "raspisaniya/move_students.html", {
        "from_group": from_group,
        "to_group": to_group,
        "students": from_group.students.all(),
    })


@login_required
def delete_unscheduled_group(request, pk):
    group = get_object_or_404(CourseGroup, pk=pk, is_scheduled=False)
    if request.method == "POST":
        subject = group.course.subject
        for st in group.students.all():
            st.debts.add(subject)
        group.delete()
        messages.success(request, "Guruh o'chirildi, talabalar qayta ro'yxatga qaytdi.")
    return redirect("build_schedule")


@login_required
def course_update(request, pk):
    course = get_object_or_404(Course, pk=pk)
    if request.method == "POST":
        start_date_raw = request.POST.get("start_date")
        end_date_raw = request.POST.get("end_date")
        total_lessons = request.POST.get("total_lessons")
        lessons_per_week = request.POST.get("lessons_per_week")

        if not all([start_date_raw, end_date_raw, total_lessons, lessons_per_week]):
            messages.error(request, "Barcha maydonlarni to'ldiring")
            return redirect("course_update", pk=pk)

        course.start_date = parse_date(start_date_raw)
        course.end_date = parse_date(end_date_raw)
        course.total_lessons = int(total_lessons)
        lessons_per_week = int(lessons_per_week)

        # ── YANGI TEKSHIRUV: jadval tuzuvchi algoritm total_lessons turiga
        # (24/16/8 paralik) qarab HAR DOIM qat'iy haftalik dars sonini
        # qo'llaydi — "Haftada necha marta" maydoniga boshqa son kiritilgan
        # bo'lsa ham, haqiqiy jadval baribir shu qat'iy songa muvofiq
        # tuziladi. Shuning uchun bu yerda mos qiymatga to'g'irlab, adminni
        # ogohlantiramiz — aks holda u "Haftada necha marta"da ko'rgan soni
        # bilan haqiqiy natija mos kelmay, keyin tushunarsiz bo'lib qolardi ──
        expected_per_week = get_expected_lessons_per_week(course.total_lessons)
        if lessons_per_week != expected_per_week:
            messages.warning(
                request,
                f"⚠ Diqqat: {course.total_lessons} paralik kurs uchun tizim haftada "
                f"faqat {expected_per_week} para joylashtira oladi (siz {lessons_per_week} "
                f"kiritgan edingiz). Qiymat avtomatik {expected_per_week} ga to'g'irlandi."
            )
        course.lessons_per_week = expected_per_week
        course.save()

        course.groups.update(is_scheduled=False)
        GroupSchedule.objects.filter(group__course=course).delete()

        messages.success(request, "Kurs yangilandi! Qayta jadval tuzing.")
        return redirect("lesson_list")

    return render(request, "raspisaniya/course_update.html", {"course": course})


# ─────────────────────────────────────────
# HAFTALIK JADVAL
# ─────────────────────────────────────────
@login_required
def weekly_schedule_view(request):
    week_str = request.GET.get('week')
    if week_str:
        try:
            week_start = dt_date.fromisoformat(week_str)
            week_start = week_start - timedelta(days=week_start.weekday())
        except Exception:
            week_start = None
    else:
        week_start = None

    data = get_weekly_schedule_data(week_start)
    grid = data['grid']
    week_start = data['week_start']
    week_end = data['week_end']
    max_group = data['max_group']

    prev_week = (week_start - timedelta(weeks=1)).isoformat()
    next_week = (week_start + timedelta(weeks=1)).isoformat()

    group_numbers = list(range(1, max_group + 1))

    SUBJECT_COLORS = [
        {'bg': '#dbeafe', 'text': '#1e40af', 'border': '#93c5fd'},
        {'bg': '#d1fae5', 'text': '#065f46', 'border': '#6ee7b7'},
        {'bg': '#fef3c7', 'text': '#92400e', 'border': '#fcd34d'},
        {'bg': '#fce7f3', 'text': '#9d174d', 'border': '#f9a8d4'},
        {'bg': '#ede9fe', 'text': '#5b21b6', 'border': '#c4b5fd'},
        {'bg': '#ffedd5', 'text': '#9a3412', 'border': '#fdba74'},
        {'bg': '#cffafe', 'text': '#155e75', 'border': '#67e8f9'},
        {'bg': '#dcfce7', 'text': '#14532d', 'border': '#86efac'},
        {'bg': '#fee2e2', 'text': '#991b1b', 'border': '#fca5a5'},
        {'bg': '#f0fdf4', 'text': '#166534', 'border': '#bbf7d0'},
        {'bg': '#fdf4ff', 'text': '#6b21a8', 'border': '#e879f9'},
        {'bg': '#fff7ed', 'text': '#9a3412', 'border': '#fed7aa'},
    ]
    subject_color_map = {}
    color_idx = [0]

    def get_subject_color(subject_name):
        if subject_name not in subject_color_map:
            subject_color_map[subject_name] = SUBJECT_COLORS[color_idx[0] % len(SUBJECT_COLORS)]
            color_idx[0] += 1
        return subject_color_map[subject_name]

    table_data = []
    for day_idx, day_name in enumerate(WEEKDAY_LIST):
        for para_idx, (start, end) in enumerate(PARA_TIMES_WEEKLY):
            cells = []
            has_any = False
            for gnum in group_numbers:
                key = (day_idx, para_idx, gnum)
                info = grid.get(key)
                if info:
                    has_any = True
                    color = get_subject_color(info['subject'])
                    cells.append({
                        'filled': True,
                        'sched_id': info['sched_id'],
                        'subject': info['subject'],
                        'teacher': info['teacher'],
                        'room': info.get('room', ''),
                        'group_number': info.get('group_number', ''),
                        'bg': color['bg'],
                        'text': color['text'],
                        'border': color['border'],
                    })
                else:
                    cells.append({'filled': False})
            table_data.append({
                'day': day_name,
                'time': f"{start} - {end}",
                'iso_date': (week_start + timedelta(days=day_idx)).isoformat(),
                'start_time': start,
                'cells': cells,
                'has_any': has_any,
                'show_day': para_idx == 0,
                'para_count': len(PARA_TIMES_WEEKLY),
            })

    return render(request, "raspisaniya/weekly_schedule.html", {
        "group_numbers": group_numbers,
        "table_data": table_data,
        "week_start": week_start,
        "week_end": week_end,
        "week_start_str": week_start.strftime("%d.%m.%Y"),
        "week_end_str": week_end.strftime("%d.%m.%Y"),
        "prev_week": prev_week,
        "next_week": next_week,
    })


@login_required
def weekly_schedule_excel(request):
    week_str = request.GET.get('week')
    if week_str:
        try:
            week_start = dt_date.fromisoformat(week_str)
            week_start = week_start - timedelta(days=week_start.weekday())
        except Exception:
            week_start = None
    else:
        week_start = None

    data = get_weekly_schedule_data(week_start)
    max_group = data['max_group']
    grid = data['grid']
    group_numbers = list(range(1, max_group + 1))

    wb = Workbook()
    ws = wb.active
    ws.title = "Haftalik jadval"

    thin = Side(style='thin', color='BBBBBB')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal='center', vertical='center', wrap_text=True)
    header_font = Font(name='Arial', bold=True, color='FFFFFF', size=10)
    header_fill = PatternFill('solid', start_color='2E4053')
    time_fill = PatternFill('solid', start_color='5D6D7E')
    time_font = Font(name='Arial', bold=True, color='FFFFFF', size=9)
    day_fill = PatternFill('solid', start_color='1A252F')
    day_font = Font(name='Arial', bold=True, color='FFFFFF', size=10)
    empty_fill = PatternFill('solid', start_color='F5F5F5')

    CELL_COLORS = [
        "D6E4BC", "B8D4E8", "FCE4A8", "E8C8D4",
        "CCE8CC", "FFD8B0", "D8D0E8", "E8E8C8",
        "BCE4E4", "FFC8C8", "D4E4F4", "E4D4BC",
        "C8D8F4", "F4D4C8", "D4F4D4", "F4F4C8",
    ]

    ws.column_dimensions['A'].width = 13
    ws.column_dimensions['B'].width = 14
    for i in range(len(group_numbers)):
        ws.column_dimensions[get_column_letter(i + 3)].width = 22

    ws.row_dimensions[1].height = 30
    for col, val in enumerate(["Kun", "Vaqt"], 1):
        c = ws.cell(1, col, val)
        c.font = header_font
        c.fill = header_fill
        c.alignment = center
        c.border = border

    for i, gnum in enumerate(group_numbers):
        c = ws.cell(1, i + 3, f"{gnum}-guruh")
        c.font = header_font
        c.fill = header_fill
        c.alignment = center
        c.border = border

    subject_color_map = {}
    color_counter = [0]

    def get_subject_color(subj):
        if subj not in subject_color_map:
            subject_color_map[subj] = CELL_COLORS[color_counter[0] % len(CELL_COLORS)]
            color_counter[0] += 1
        return subject_color_map[subj]

    row = 2
    for day_idx, day_name in enumerate(WEEKDAY_LIST):
        day_start_row = row
        for para_idx, (start, end) in enumerate(PARA_TIMES_WEEKLY):
            ws.row_dimensions[row].height = 45

            tc = ws.cell(row, 2, f"{start} - {end}")
            tc.font = time_font
            tc.fill = time_fill
            tc.alignment = center
            tc.border = border

            for i, gnum in enumerate(group_numbers):
                col = i + 3
                key = (day_idx, para_idx, gnum)
                cell = ws.cell(row, col)
                info = grid.get(key)
                if info:
                    room_str = f"\n🏫 {info['room']}" if info.get('room') else ''
                    group_str = f" ({info['group_number']}-guruh)" if info.get('group_number') else ''
                    cell.value = f"{info['subject']}{group_str}\n{info['teacher']}{room_str}"
                    cell.fill = PatternFill('solid', start_color=get_subject_color(info['subject']))
                    cell.font = Font(name='Arial', size=9, bold=True)
                else:
                    cell.value = ""
                    cell.fill = empty_fill
                    cell.font = Font(name='Arial', size=9)
                cell.alignment = center
                cell.border = border

            row += 1

        if row - day_start_row > 1:
            ws.merge_cells(
                start_row=day_start_row, start_column=1,
                end_row=row - 1, end_column=1
            )
        kc = ws.cell(day_start_row, 1)
        kc.value = day_name
        kc.font = day_font
        kc.fill = day_fill
        kc.alignment = center
        kc.border = border

    ws.freeze_panes = 'C2'

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = 'attachment; filename="haftalik_jadval.xlsx"'
    wb.save(response)
    return response

@login_required
def change_lesson_time_ajax(request, sched_pk):
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'Faqat POST so\'rov'}, status=405)

    try:
        body = json.loads(request.body)
    except (json.JSONDecodeError, ValueError):
        return JsonResponse({'success': False, 'error': 'JSON xato'}, status=400)

    sched = get_object_or_404(GroupSchedule, pk=sched_pk)

    new_date_raw = body.get('new_date', '').strip()
    new_time_raw = body.get('new_time', '').strip()

    if not new_date_raw or not new_time_raw:
        return JsonResponse({'success': False, 'error': 'Sana yoki vaqt yuborilmagan'}, status=400)

    new_date_val = parse_date(new_date_raw)
    if not new_date_val:
        return JsonResponse({'success': False, 'error': 'Noto\'g\'ri sana formati'}, status=400)

    try:
        h, m = map(int, new_time_raw.split(':'))
        new_time_val = dtime(h, m)
    except (ValueError, AttributeError):
        return JsonResponse({'success': False, 'error': 'Noto\'g\'ri vaqt formati'}, status=400)

    if sched.date == new_date_val and sched.start_time == new_time_val:
        return JsonResponse({'success': False, 'error': 'Dars allaqachon shu vaqtda'})

    # Faqat bugungi kun o'zgartiriladi (admin istisno yoki ruxsat berilgan)
    from datetime import date as dt_date_check
    today = dt_date_check.today()
    is_admin = request.user.is_superuser or request.user.is_staff

    if not is_admin:
        if sched.date == today:
            # Bugungi kun — o'zgartirish mumkin, ruxsat shart emas
            pass
        elif sched.group.teacher_can_edit:
            # Admin ruxsat bergan — o'zgartirish mumkin
            pass
        else:
            # Boshqa kun + ruxsat yo'q — bloklash
            return JsonResponse({
                'success': False,
                'error': f'Faqat bugungi ({today.strftime("%d.%m.%Y")}) darsni o\'zgartirish mumkin! Boshqa kunlar uchun admin ruxsat berishi kerak.'
            })
    # ── YANGI: "kelajakdagi barcha haftalarga qo'llash" — faqat ADMIN uchun
    # (o'qituvchi bugungi bitta darsni o'zgartirish huquqiga ega, lekin butun
    # kelajakdagi jadvalni ommaviy o'zgartirish faqat admin qo'lida bo'lishi
    # kerak) ──
    apply_to_future = bool(body.get('apply_to_future')) and is_admin

    updated, skipped_dates = apply_lesson_time_change(
        sched, new_date_val, new_time_val, apply_to_future=apply_to_future
    )

    if updated == 0:
        if skipped_dates:
            return JsonResponse({
                'success': False,
                'error': f'{skipped_dates[0].strftime("%d.%m.%Y")} kuni o\'qituvchi yoki talabalar band!'
            })
        return JsonResponse({'success': False, 'error': 'Dars topilmadi yoki o\'zgartirilmadi'})

    end_time = (datetime.combine(new_date_val, new_time_val) + timedelta(minutes=80)).time()

    return JsonResponse({
        'success':        True,
        'new_date':       new_date_val.strftime('%d.%m.%Y'),
        'new_date_iso':   new_date_val.isoformat(),
        'new_time':       new_time_val.strftime('%H:%M'),
        'end_time':       end_time.strftime('%H:%M'),
        'weekday':        WEEKDAY_NAMES.get(new_date_val.weekday(), ''),
        'updated_count':  updated,
        'skipped_count':  len(skipped_dates),
    })



@login_required
def toggle_teacher_edit_permission(request, group_pk):
    """Admin: guruh uchun o'qituvchiga dars vaqtini o'zgartirish ruxsatini berish/olish."""
    if not (request.user.is_staff or request.user.is_superuser):
        return JsonResponse({'success': False, 'error': 'Ruxsat yo\'q'}, status=403)
    group = get_object_or_404(CourseGroup, pk=group_pk)
    if request.method == 'POST':
        group.teacher_can_edit = not group.teacher_can_edit
        group.save(update_fields=['teacher_can_edit'])
        return JsonResponse({
            'success': True,
            'permitted': group.teacher_can_edit,
            'label': 'Ruxsat berildi ✅' if group.teacher_can_edit else 'Ruxsat berilmagan',
        })
    return JsonResponse({'success': False}, status=405)

@staff_member_required
def reset_database_view(request):
    if request.method == 'POST' and request.POST.get('confirm') == 'TASDIQLASH':
        selected_models = request.POST.getlist('models_to_delete')

        if not selected_models:
            return render(request, 'raspisaniya/reset_database.html', {
                'error': "Hech bo'lmaganda bitta bo'limni tanlang!",
                'done': False
            })

        try:
            with transaction.atomic():
                # ── 1. Avval user ID larini yig'ib ol (o'chirishdan oldin) ──
                student_user_ids = []
                teacher_user_ids = []

                if 'student' in selected_models:
                    student_user_ids = list(
                        Student.objects.filter(user__isnull=False)
                        .values_list('user_id', flat=True)
                    )
                if 'teacher' in selected_models:
                    teacher_user_ids = list(
                        Teacher.objects.filter(user__isnull=False)
                        .values_list('user_id', flat=True)
                    )

                # ── 2. To'g'ri tartibda o'chirish (CASCADE yo'q — faqat Django ORM) ──
                # Tartib muhim: avval bog'liq jadvallar, keyin asosiylar

                if 'schedule' in selected_models:
                    GroupSchedule.objects.all().delete()

                if 'group' in selected_models:
                    CourseGroup.objects.all().delete()

                if 'course' in selected_models:
                    # GroupSchedule va CourseGroup avval o'chadi (yuqorida yoki cascade orqali)
                    Course.objects.all().delete()

                if 'student' in selected_models:
                    Student.objects.all().delete()

                if 'teacher' in selected_models:
                    Teacher.objects.all().delete()

                if 'subject' in selected_models:
                    Subject.objects.all().delete()

                if 'room' in selected_models:
                    Room.objects.all().delete()

                # ── 3. Tegishli oddiy foydalanuvchilarni o'chir ──
                all_user_ids = list(set(student_user_ids + teacher_user_ids))
                if all_user_ids:
                    User.objects.filter(
                        id__in=all_user_ids,
                        is_staff=False,
                        is_superuser=False,
                    ).delete()

            return render(request, 'raspisaniya/reset_database.html', {'done': True})

        except Exception as e:
            print(f"Baza tozalashda xato: {e}")
            return render(request, 'raspisaniya/reset_database.html', {
                'error': f"Xatolik yuz berdi: {str(e)}",
                'done': False
            })

    return render(request, 'raspisaniya/reset_database.html', {'done': False})


def get_backups_dir():
    backup_dir = os.path.join(settings.MEDIA_ROOT, 'backups')
    os.makedirs(backup_dir, exist_ok=True)
    return backup_dir


def export_database_view(request):
    """Backup yaratish — serverda saqlaydi va yuklab olish imkonini beradi."""
    from datetime import datetime as dt

    output = StringIO()
    management.call_command(
        'dumpdata',
        'raspisaniya', 'accounts', 'auth.user',
        indent=2,
        stdout=output,
        natural_foreign=True,
        natural_primary=True,
    )
    data = output.getvalue()

    # Nom: foydalanuvchi bergan nom yoki avtomatik sana
    custom_name = request.GET.get('name', '').strip()
    now = dt.now().strftime('%Y-%m-%d_%H-%M')
    if custom_name:
        safe_name = ''.join(c for c in custom_name if c.isalnum() or c in '-_ ')
        safe_name = safe_name.strip().replace(' ', '_')
        filename = f"{now}_{safe_name}.json"
    else:
        filename = f"{now}_backup.json"

    # Serverda saqlash
    backup_dir = get_backups_dir()
    filepath = os.path.join(backup_dir, filename)
    with open(filepath, 'w', encoding='utf-8') as f:
        f.write(data)

    # Yuklab olish
    response = HttpResponse(data, content_type='application/json')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


def restore_database_view(request):
    backup_dir = get_backups_dir()

    def do_restore(filepath):
        """Bazani tiklash — avval eski ma'lumotlarni tozalab, keyin yuklaymiz."""
        from django.contrib.auth.models import User
        from raspisaniya.models import (
            Student, Teacher, Subject, CourseGroup, Group,
            GroupSchedule, Room, Course, Attendance, Grade
        )
        try:
            # Ketma-ketlikda tozalaymiz (foreign key tartibida)
            Attendance.objects.all().delete()
            Grade.objects.all().delete()
            GroupSchedule.objects.all().delete()
            CourseGroup.objects.all().delete()
            Course.objects.all().delete()
            Student.objects.all().delete()
            Teacher.objects.all().delete()
            Room.objects.all().delete()
            Subject.objects.all().delete()
            # MUHIM: `Group` (talaba guruhi, masalan "21-KI-01") oldin o'chirilmagan edi —
            # shu sabab eski nomlar bazada qolib, tiklashda `UNIQUE constraint failed:
            # raspisaniya_group.name` xatosini berardi.
            Group.objects.all().delete()
            # Auth userlarni ham tozalaymiz (superuser qolsin)
            User.objects.filter(is_superuser=False).delete()
        except Exception as e:
            return f"Tozalashda xato: {e}"

        try:
            management.call_command('loaddata', filepath, ignorenonexistent=True)
            return None  # Xato yo'q
        except Exception as e:
            return str(e)

    if request.method == 'POST':
        action = request.POST.get('action')

        # 1. Fayl yuklash orqali tiklash
        if action == 'upload' and request.FILES.get('backup_file'):
            backup_file = request.FILES['backup_file']
            custom_name = request.POST.get('backup_name', '').strip()
            from datetime import datetime as dt
            now = dt.now().strftime('%Y-%m-%d_%H-%M')

            if custom_name:
                safe_name = ''.join(c for c in custom_name if c.isalnum() or c in '-_ ')
                filename = f"{now}_{safe_name.replace(' ', '_')}.json"
            else:
                filename = f"{now}_{backup_file.name}"

            save_path = os.path.join(backup_dir, filename)
            with open(save_path, 'wb+') as dest:
                for chunk in backup_file.chunks():
                    dest.write(chunk)

            err = do_restore(save_path)
            if err:
                messages.error(request, f"Xatolik: {err}")
            else:
                messages.success(request, f"✅ '{filename}' yuklandi va baza tiklandi!")
            return redirect('restore_database')

        # 2. Serverda saqlangan backup dan tiklash
        if action == 'restore_saved':
            filename = request.POST.get('filename', '')
            filepath = os.path.join(backup_dir, filename)
            if os.path.exists(filepath):
                err = do_restore(filepath)
                if err:
                    messages.error(request, f"Xatolik: {err}")
                else:
                    messages.success(request, f"✅ '{filename}' dan baza muvaffaqiyatli tiklandi!")
            else:
                messages.error(request, "Fayl topilmadi!")
            return redirect('restore_database')

        # 3. Backup o'chirish
        if action == 'delete_backup':
            filename = request.POST.get('filename', '')
            filepath = os.path.join(backup_dir, filename)
            if os.path.exists(filepath) and filename.endswith('.json'):
                os.remove(filepath)
                messages.success(request, f"'{filename}' o'chirildi.")
            return redirect('restore_database')

    # Saqlangan backuplar ro'yxati
    backups = []
    if os.path.exists(backup_dir):
        for fname in sorted(os.listdir(backup_dir), reverse=True):
            if fname.endswith('.json'):
                fpath = os.path.join(backup_dir, fname)
                stat = os.stat(fpath)
                backups.append({
                    'name': fname,
                    'size': round(stat.st_size / 1024, 1),
                    'date': stat.st_mtime,
                })

    return render(request, 'raspisaniya/restore_database.html', {
        'backups': backups,
    })


@login_required
def admin_change_teacher_password(request, pk):
    """Admin: o'qituvchi parolini o'zgartirish."""
    teacher = get_object_or_404(Teacher, pk=pk)
    if request.method == "POST":
        new_password = request.POST.get("new_password", "").strip()
        if not new_password:
            messages.error(request, "Parol bo'sh bo'lmasin")
        elif not teacher.user:
            messages.error(request, "Bu o'qituvchining tizim akkaunti yo'q")
        else:
            teacher.user.set_password(new_password)
            teacher.user.save()
            messages.success(request, f"{teacher.first_name} ning paroli o'zgartirildi")
    return redirect("teacher_list")


@login_required
def toggle_all_teacher_edit_permission(request):
    """Admin: barcha guruhlarga bir vaqtda ruxsat berish/olish."""
    if not (request.user.is_staff or request.user.is_superuser):
        return JsonResponse({'success': False, 'error': 'Ruxsat yo\'q'}, status=403)
    if request.method == 'POST':
        import json as _json
        body = _json.loads(request.body)
        permitted = body.get('permitted', True)
        CourseGroup.objects.filter(is_scheduled=True).update(teacher_can_edit=permitted)
        count = CourseGroup.objects.filter(is_scheduled=True).count()
        return JsonResponse({
            'success': True,
            'permitted': permitted,
            'count': count,
            'label': f'Hammaga ruxsat berildi ✅ ({count} guruh)' if permitted else f'Hammadan ruxsat olindi ({count} guruh)',
        })
    return JsonResponse({'success': False}, status=405)


@login_required
def sched_info(request, sched_pk):
    """Dars haqida talabalar ro'yxati — modal uchun JSON."""
    sched = get_object_or_404(GroupSchedule, pk=sched_pk)
    students = list(
        sched.group.students.all()
        .order_by('first_name')
        .values_list('first_name', flat=True)
    )
    return JsonResponse({
        'students': students,
        'total': len(students),
        'subject': str(sched.group.course.subject),
        'teacher': str(sched.group.teacher),
        'room': str(sched.group.room) if sched.group.room else '—',
        'group_number': sched.group.group_number,
    })


def student_schedule_info(request, student_pk):
    try:
        # Talabani bazadan qidiramiz
        student = get_object_or_404(Student, pk=student_pk)

        WEEKDAY_NAMES_LOCAL = {
            0: 'Dushanba', 1: 'Seshanba', 2: 'Chorshanba',
            3: 'Payshanba', 4: 'Juma', 5: 'Shanba', 6: 'Yakshanba'
        }

        # Talaba a'zo bo'lgan barcha guruhlarni (CourseGroup) yuklaymiz
        groups = CourseGroup.objects.filter(
            students=student
        ).select_related('course__subject', 'teacher', 'room').prefetch_related('schedule')

        result = []
        for grp in groups:
            days_set = {}
            # Guruhning bazadagi barcha dars sanalarini tekshiramiz
            all_scheds = grp.schedule.all().order_by('date', 'start_time')

            for sched in all_scheds:
                st = sched.start_time or grp.start_time
                if not st:
                    continue

                wd = sched.date.weekday()
                key = (wd, st.strftime('%H:%M'))

                if key not in days_set:
                    end_min = st.hour * 60 + st.minute + 80
                    end_h, end_m = divmod(end_min, 60)
                    days_set[key] = {
                        'weekday': WEEKDAY_NAMES_LOCAL.get(wd, ''),
                        'time': f"{st.strftime('%H:%M')} – {end_h:02d}:{end_m:02d}",
                    }

            # Hafta kunlari tartibida saralash
            days = sorted(
                days_set.values(),
                key=lambda d: list(WEEKDAY_NAMES_LOCAL.values()).index(d['weekday']) if d[
                                                                                            'weekday'] in WEEKDAY_NAMES_LOCAL.values() else 9
            )

            result.append({
                'subject': str(grp.course.subject),
                'teacher': f"{grp.teacher.last_name} {grp.teacher.first_name}".strip() if grp.teacher else "O'qituvchi",
                'room': str(grp.room) if grp.room else '—',
                'days': days,
            })

        return JsonResponse({'success': True, 'groups': result})

    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=500)


# ==============================================================================
# 2. HAFTALIK UMUMIY JADVALDA DARS KARTASI BOSILGANDA (sched_id orqali) ISHLAYDIGAN FUNKSIYA
# ==============================================================================
def sched_info_ajax(request, sched_id):
    try:
        current_sched = get_object_or_404(GroupSchedule, pk=sched_id)
        course_group = current_sched.group

        WEEKDAY_NAMES_LOCAL = {
            0: 'Dushanba', 1: 'Seshanba', 2: 'Chorshanba',
            3: 'Payshanba', 4: 'Juma', 5: 'Shanba', 6: 'Yakshanba'
        }

        # Talabalar ro'yxati
        students_list = []
        if course_group:
            for student in course_group.students.all():
                full_name = f"{student.last_name} {student.first_name}".strip()
                if not full_name:
                    full_name = student.user.get_full_name() if student.user else str(student)

                lang_code = student.language or course_group.language or 'uz'
                student_lang = 'RUS' if lang_code.lower() == 'ru' else 'UZB'

                students_list.append({
                    'name': full_name,
                    'lang': student_lang
                })
        students_list = sorted(students_list, key=lambda x: x['name'])

        # Guruhning dars jadvali shablonini (barcha kunlarini) aniqlash
        other_days_result = []
        if course_group:
            # Guruhning barcha dars kunlarini yig'amiz
            all_schedules = GroupSchedule.objects.filter(group=course_group).order_by('date')
            days_set = {}

            for sched in all_schedules:
                st = sched.start_time or course_group.start_time
                if not st:
                    continue

                wd = sched.date.weekday()
                key = (wd, st.strftime('%H:%M'))

                if key not in days_set:
                    end_min = st.hour * 60 + st.minute + 80
                    end_h, end_m = divmod(end_min, 60)

                    days_set[key] = {
                        'weekday': WEEKDAY_NAMES_LOCAL.get(wd, ''),
                        'time': f"{st.strftime('%H:%M')} – {end_h:02d}:{end_m:02d}",
                    }

            sorted_days = sorted(
                days_set.values(),
                key=lambda d: list(WEEKDAY_NAMES_LOCAL.values()).index(d['weekday']) if d[
                                                                                            'weekday'] in WEEKDAY_NAMES_LOCAL.values() else 9
            )
            other_days_result = [f"{d['weekday']} ({d['time']})" for d in sorted_days]

        return JsonResponse({
            'success': True,
            'students': students_list,
            'other_days': other_days_result
        })

    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=500)


class NumberedCanvas(canvas.Canvas):
    """PDF sahifalarining ostiga 'Sahifa X / Y' dinamik raqamini qo'yish uchun"""

    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        self._saved_page_states = []

    def showPage(self):
        self._saved_page_states.append(dict(self.__dict__))
        self._startPage()

    def save(self):
        num_pages = len(self._saved_page_states)
        for state in self._saved_page_states:
            self.__dict__.update(state)
            self.draw_page_number(num_pages)
            super().showPage()
        super().save()

    def draw_page_number(self, page_count):
        self.setFont("Helvetica", 8)
        self.setFillColor(colors.HexColor("#555555"))
        page_text = f"Sahifa {self._pageNumber} / {page_count}"
        self.drawRightString(A4[0] - 30, 20, page_text)


@login_required
def download_vedomost(request, group_id):
    # SIKLIK IMPORT (Circular Import) xatoligini oldini olish uchun importni funksiya ichida bajaramiz
    from .models import CourseGroup, Attendance, Grade
    import io
    import datetime
    from django.http import HttpResponse
    from reportlab.lib.pagesizes import A4
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib import colors

    # Guruh ma'lumotlarini bazadan olish
    group = get_object_or_404(CourseGroup, pk=group_id)
    students = group.students.all().order_by('last_name', 'first_name')

    # 1. AVTOMATIK VEDOMOST RAQAMINI SHAKLLANTIRISH
    start_year = group.course.start_date.year if group.course.start_date else datetime.date.today().year
    end_year = start_year + 1
    oq_yil = f"{start_year}/{end_year}"

    qayta_oqish_status = "1"
    vedomost_no = f"{oq_yil}/{qayta_oqish_status}-{group.pk}"

    # Qaydnoma to'ldirilgan sana
    qayd_sana = datetime.date.today().strftime("%d.%m.%Y")

    # PDF sahifasi o'lchamlari (A4)
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=A4,
        rightMargin=30,
        leftMargin=30,
        topMargin=30,
        bottomMargin=40
    )

    styles = getSampleStyleSheet()

    # Matn va sarlavhalar stillari
    title_style = ParagraphStyle(
        'VTitle', parent=styles['Normal'], fontName='Helvetica-Bold', fontSize=12, leading=16, alignment=1, spaceAfter=4
    )
    subtitle_style = ParagraphStyle(
        'VSub', parent=styles['Normal'], fontName='Helvetica', fontSize=9, leading=14, alignment=0, spaceAfter=4
    )

    # Oddiy matnlar stili (F.I.Sh va Haqiqiy guruh uchun - chapga tekislangan)
    table_text_style = ParagraphStyle(
        'VText', parent=styles['Normal'], fontName='Helvetica', fontSize=8, leading=10
    )

    # Raqamlar va baholar stili (No, JN, ON, YN, Jami, Baho uchun - o'rtaga tekislangan)
    table_center_text = ParagraphStyle(
        'VCenterText', parent=styles['Normal'], fontName='Helvetica', fontSize=8, leading=10, alignment=1
    )

    # Jadval sarlavhasi (Header) stili
    table_header_style = ParagraphStyle(
        'VHeader', parent=styles['Normal'], fontName='Helvetica-Bold', fontSize=8, leading=10, alignment=1,
        textColor=colors.whitesmoke
    )

    elements = []

    # Sarlavha qismini shakllantirish
    elements.append(Paragraph("TOSHKENT FARMATSEVTIKA INSTITUTI", title_style))
    elements.append(Paragraph(f"BAHOLASH QAYDNOMASI № {vedomost_no}", title_style))
    elements.append(Spacer(1, 10))

    # Hujjat haqida avtomatik ma'lumotlar
    elements.append(Paragraph(f"<b>Fan nomi:</b> {group.course.subject.name}", subtitle_style))
    elements.append(
        Paragraph(f"<b>Fan o'qituvchisi:</b> {group.teacher.last_name} {group.teacher.first_name}", subtitle_style))
    elements.append(Paragraph(f"<b>Qaydnoma to'ldirilgan sana:</b> {qayd_sana}", subtitle_style))
    elements.append(Spacer(1, 12))

    # Jadval ustunlari sarlavhasi
    headers = [
        Paragraph("<b>No</b>", table_header_style),
        Paragraph("<b>Talabaning familiyasi, ismi, sharifi</b>", table_header_style),
        Paragraph("<b>Guruhi</b>", table_header_style),
        Paragraph("<b>JN</b><br/><font size=6>max 30</font>", table_header_style),
        Paragraph("<b>ON</b><br/><font size=6>max 20</font>", table_header_style),
        Paragraph("<b>YN</b><br/><font size=6>max 50</font>", table_header_style),
        Paragraph("<b>Umumiy ball</b><br/><font size=6>max 100</font>", table_header_style),
        Paragraph("<b>Baho</b>", table_header_style),
        Paragraph("<b>O'qituvchi imzosi</b>", table_header_style),
    ]

    table_data = [headers]

    # Talabalarni aylantirib jadval qatorlarini to'ldirish
    for idx, student in enumerate(students, 1):
        full_name = f"{student.last_name} {student.first_name}"
        haqiqiy_guruh = student.group.name if student.group else "Mavjud emas"

        # Baholarni Grade modelidan olish
        grade_obj = Grade.objects.filter(student=student, course_group=group).first()

        # Davomat foizini aniqlash (25% lik blok sharti uchun)
        total_lessons = group.schedule.count()
        missed_count = Attendance.objects.filter(student=student, schedule__in=group.schedule.all(),
                                                 is_present=False).count()
        missed_percent = (missed_count / total_lessons * 100) if total_lessons > 0 else 0

        # 🌟 O'ZGARISH SHU YERDA: Bloklangan yoki yiqilgan talabaga 2 qo'yilmaydi
        if missed_percent > 25:
            jn = "0"
            on = "0"
            yn = "0"
            umumiy = "0"
            baho = "-"  # "2 (Blok)" o'rniga faqat chiziqcha qoldiramiz
        else:
            jn_val = grade_obj.current if (grade_obj and grade_obj.current is not None) else 0
            on_val = grade_obj.midterm if (grade_obj and grade_obj.midterm is not None) else 0
            yn_val = grade_obj.final if (grade_obj and grade_obj.final is not None) else 0

            total_val = jn_val + on_val + yn_val

            jn = str(jn_val)
            on = str(on_val)
            yn = str(yn_val)
            umumiy = str(total_val)

            # Reyting shkalasi konvertatsiyasi
            if total_val >= 86:
                baho = "5"
            elif total_val >= 71:
                baho = "4"
            elif total_val >= 56:
                baho = "3"
            else:
                baho = "-"  # 56 dan kam bo'lsa ham "2" qo'yilmaydi, bo'sh (chiziqcha) qoladi

        # Qator ma'lumotlarini o'z stillari bilan jadvalga qo'shish
        table_data.append([
            Paragraph(str(idx), table_center_text),
            Paragraph(f"<b>{full_name}</b>", table_text_style),
            Paragraph(haqiqiy_guruh, table_text_style),
            Paragraph(jn, table_center_text),
            Paragraph(on, table_center_text),
            Paragraph(yn, table_center_text),
            Paragraph(f"<b>{umumiy}</b>", table_center_text),
            Paragraph(baho, table_center_text),
            Paragraph("", table_text_style),
        ])

    # Ustunlar kengligi
    col_widths = [28, 155, 72, 35, 35, 35, 50, 60, 65]

    vedomost_table = Table(table_data, colWidths=col_widths, repeatRows=1)

    # Jadvalning vizual stillari
    vedomost_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor("#1a237e")),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor("#cccccc")),

        ('TOPPADDING', (0, 0), (-1, -1), 7),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 7),

        ('ALIGN', (1, 1), (1, -1), 'LEFT'),
        ('ALIGN', (2, 1), (2, -1), 'LEFT'),
    ]))

    elements.append(vedomost_table)
    elements.append(Spacer(1, 15))

    # Jami talabalar soni qismi
    elements.append(Paragraph(f"<b>Jami talabalar soni:</b> {len(students)} ta", subtitle_style))
    elements.append(Spacer(1, 20))

    # Imzolar qismi
    signature_data = [
        [Paragraph("<b>Fakultet dekani:</b> ___________________________", subtitle_style),
         Paragraph("<b>Kafedra mudiri:</b> ___________________________", subtitle_style)],
        [Spacer(1, 15), Spacer(1, 15)],
    ]
    signature_table = Table(signature_data, colWidths=[265, 265])
    elements.append(signature_table)

    # PDF faylni qurish (NumberedCanvas ishlashi uchun u loyihada sozlangan bo'lishi kerak)
    try:
        from .utils import NumberedCanvas  # Yoki NumberedCanvas qayerda yozilgan bo'lsa o'sha yerdan import qilinadi
        doc.build(elements, canvasmaker=NumberedCanvas)
    except Exception:
        doc.build(elements)  # Agar muammo bo'lsa oddiy build qiladi

    buffer.seek(0)

    # Brauzerga yuklash uchun javob yuborish
    response = HttpResponse(buffer, content_type='application/pdf')
    response[
        'Content-Disposition'] = f'attachment; filename="Qaydnoma_{group.course.subject.name}_{group.group_number}-guruh.pdf"'
    return response


def admin_password_reset_request(request):
    """ Admin parolini tiklash: Kod har doim loyiha rahbarining pochtasiga boradi """
    if request.method == "POST":
        # Formadan kelayotgan emailni tekshirib o'tirmaymiz,
        # chunki kod baribir settings.py dagi sizning emailingizga ketadi.

        # Tizimdagi asosiy superuser (admin)ni topamiz
        admin_user = User.objects.filter(is_superuser=True).first()

        if admin_user:
            # 5 xonali tasodifiy kod yaratish
            code = str(random.randint(10000, 99999))

            # Kod va admin ID sini sessionga saqlaymiz
            request.session['reset_code'] = code
            request.session['reset_admin_id'] = admin_user.pk

            # Email sozlamalari
            subject = "Dars jadvali - Admin parolini va loginini tiklash kodi"
            message = f"Tizim admin paneli uchun parolni tiklash kodi: {code}\n\nUshbu kodni parolni yangilash sahifasiga kiriting."

            # Kod yuboriladigan manzil - qat'iy ravishda settings.py dagi sizning pochtangiz
            target_email = settings.ADMIN_RESET_TARGET_EMAIL
            from_email = settings.DEFAULT_FROM_EMAIL

            try:
                send_mail(subject, message, from_email, [target_email], fail_silently=False)
                messages.success(request, f"Tasdiqlash kodi tizim rahbarining pochtasiga yuborildi.")
                return render(request, "accounts/admin_password_verify.html")
            except Exception as e:
                messages.error(request, f"Email yuborishda xatolik: {str(e)}")
                return redirect('login')
        else:
            messages.error(request, "Tizimda hech qanday admin foydalanuvchi topilmadi!")
            return redirect('login')

    return redirect('login')


def admin_password_verify_and_change(request):
    """ Kodni tekshirish, login (username) va parolni yangilash """
    if request.method == "POST":
        input_code = request.POST.get("verification_code", "").strip()
        new_username = request.POST.get("new_username", "").strip()
        new_password = request.POST.get("new_password", "").strip()
        confirm_password = request.POST.get("confirm_password", "").strip()

        session_code = request.session.get('reset_code')
        admin_id = request.session.get('reset_admin_id')

        if not session_code or not admin_id:
            messages.error(request, "Sessiya muddati tugagan yoki noto'g'ri so'rov.")
            return redirect('login')

        if input_code != session_code:
            messages.error(request, "Kiritilgan tasdiqlash kodi noto'g'ri!")
            return render(request, "accounts/admin_password_verify.html")

        # 🌟 YANGI LOGIN TEKSHIRUVI: Bu username bazada band emasligini tekshiramiz
        # Lekin aynan shu o'zgartirayotgan adminning o'z eski ID si bo'lsa, unga ruxsat berish kerak
        username_exists = User.objects.filter(username=new_username).exclude(pk=admin_id).exists()
        if username_exists:
            messages.error(request, f"'{new_username}' ID raqami tizimda band! Boshqa ID kiriting.")
            return render(request, "accounts/admin_password_verify.html")

        if new_password != confirm_password:
            messages.error(request, "Yangi parollar bir-biriga mos kelmadi.")
            return render(request, "accounts/admin_password_verify.html")

        if len(new_password) < 4:
            messages.error(request, "Parol kamida 4 ta belgi bo'lishi kerak.")
            return render(request, "accounts/admin_password_verify.html")

        # Admin modelini olib, ham username, ham parolni yangilaymiz
        admin_user = User.objects.get(pk=admin_id)
        admin_user.username = new_username
        admin_user.set_password(new_password)
        admin_user.save()

        # Sessiyadagi ma'lumotlarni tozalaymiz
        del request.session['reset_code']
        del request.session['reset_admin_id']

        messages.success(request,
                         f"Admin hisob ma'lumotlari muvaffaqiyatli yangilandi! Yangi ID ({new_username}) va yangi parol bilan kiring.")
        return redirect('login')

    return redirect('login')


# raspisaniya/views.py ga qo'shing

@login_required
def teacher_capacity_check(request):
    """
    O'qituvchilar uchun matematik imkoniyat tekshiruvi.
    Har bir o'qituvchi uchun: mavjud bo'sh paralar vs kerak bo'lgan paralar.
    """
    from django.db.models import Count, Sum

    # Jadval tuzilmagan guruhlar
    unscheduled = CourseGroup.objects.filter(
        is_scheduled=False
    ).select_related('course__subject', 'teacher', 'course').prefetch_related('students')

    # O'qituvchi bo'yicha guruhlash
    teacher_data = defaultdict(lambda: {
        'teacher': None,
        'groups': [],
        'total_needed': 0,
    })

    for grp in unscheduled:
        tid = grp.teacher_id
        teacher_data[tid]['teacher'] = grp.teacher
        teacher_data[tid]['groups'].append(grp)
        teacher_data[tid]['total_needed'] += grp.course.total_lessons

    results = []

    for tid, tdata in teacher_data.items():
        teacher = tdata['teacher']
        groups  = tdata['groups']

        # Muddatni aniqlash (eng keng muddat)
        start = min(g.course.start_date for g in groups)
        end   = max(g.course.end_date   for g in groups)

        # Ish kunlari soni
        work_days = sum(
            1 for i in range((end - start).days + 1)
            if (start + timedelta(days=i)).weekday() <= 4  # Du-Ju
        )

        # Mavjud jadvaldagi band paralar (shu muddat ichida)
        already_scheduled = GroupSchedule.objects.filter(
            group__teacher=teacher,
            date__gte=start,
            date__lte=end,
        ).count()

        # Jami mavjud joy
        total_slots = work_days * 6  # kuniga max 6 para

        # Bo'sh joy
        free_slots = total_slots - already_scheduled

        # Kerak
        needed = tdata['total_needed']

        # Imkoniyat
        possible   = free_slots >= needed
        shortage   = max(0, needed - free_slots)
        extra_days = math.ceil(shortage / 6) if shortage > 0 else 0

        # Kunlik o'rtacha yuklanma (yangi guruhlar bilan)
        avg_per_day = round((already_scheduled + needed) / max(work_days, 1), 1)

        results.append({
            'teacher':            teacher,
            'groups':             groups,
            'start':              start,
            'end':                end,
            'work_days':          work_days,
            'total_slots':        total_slots,
            'already_scheduled':  already_scheduled,
            'free_slots':         free_slots,
            'needed':             needed,
            'possible':           possible,
            'shortage':           shortage,
            'extra_days':         extra_days,
            'avg_per_day':        avg_per_day,
        })

    # Imkonsizlarni avval ko'rsatish
    results.sort(key=lambda x: (x['possible'], -x['shortage']))

    return render(request, 'raspisaniya/teacher_capacity_check.html', {
        'results': results,
        'total_impossible': sum(1 for r in results if not r['possible']),
        'total_possible':   sum(1 for r in results if r['possible']),
    })


@login_required
def assign_teachers_auto(request):
    """
    TIZIMDAGI BARCHA kurslarning o'qituvchisi yo'q guruhlariga konfliktlarsiz avtomatik o'qituvchi taqsimlash.
    """
    if request.method != "POST":
        return redirect('lesson_list')

    courses = Course.objects.filter(groups__teacher__isnull=True).distinct().select_related('subject')

    if not courses.exists():
        messages.info(request, "Tizimda o'qituvchi biriktirilmagan guruhlar topilmadi.")
        return redirect('lesson_list')

    total_assigned_count = 0
    all_failed_details = []

    for course in courses:
        groups = list(
            course.groups.filter(teacher__isnull=True)
            .prefetch_related('students', 'schedule')
        )
        candidates = list(Teacher.objects.filter(subjects=course.subject).order_by('pk'))

        if not candidates:
            for grp in groups:
                all_failed_details.append(f"{course.subject.name} ({grp.group_number}-guruh): O'qituvchi umuman yo'q")
            continue

        start = course.start_date
        end = course.end_date

        def get_teacher_free_slots_count(teacher):
            busy_count = GroupSchedule.objects.filter(
                group__teacher=teacher,
                date__gte=start,
                date__lte=end,
            ).count()
            work_days = sum(
                1 for i in range((end - start).days + 1)
                if (start + timedelta(days=i)).weekday() <= 4
            )
            return work_days * 6 - busy_count

        # Har bir guruh uchun o'qituvchi saralash
        for grp in groups:
            # MUHIM: Guruh allaqachon "Jadval tuzish" bosqichida aniq (sana, vaqt)
            # larga joylashtirilgan bo'ladi. Endi haftaning taxminiy kunlarini
            # (check_wds) TAXMIN QILISH o'rniga, guruhning HAQIQIY jadvalidan
            # foydalanamiz — bu ustozlarni bekorga rad etib, "bitta guruhdan keyin
            # to'xtab qolish" muammosini bartaraf etadi.
            grp_slots = list(grp.schedule.values_list('date', 'start_time'))

            if not grp_slots:
                all_failed_details.append(
                    f"{course.subject.name} ({grp.group_number}-guruh): "
                    f"guruh hali jadvallanmagan (avval 'Jadval tuzish'ni bajaring)"
                )
                continue

            best_teacher = None
            max_free = -1

            for teacher in candidates:
                free = get_teacher_free_slots_count(teacher)

                # 1. Yuklama yetarliligini tekshirish
                if free < course.total_lessons:
                    continue

                # 2. ANIQ TO'QNASHUV TEKSHIRUVI: ustoz aynan shu guruhning
                #    HAQIQIY (sana, vaqt) laridan birortasida allaqachon BOSHQA
                #    guruhga band bo'lsa — bu ustozni rad etamiz. Guruhning o'zi
                #    boshqa vaqt/kunda bo'lsa, ustoz erkin hisoblanadi.
                teacher_busy_slots = set(
                    GroupSchedule.objects.filter(
                        group__teacher=teacher,
                        date__gte=start,
                        date__lte=end,
                    ).exclude(group=grp).values_list('date', 'start_time')
                )
                conflict = any(slot in teacher_busy_slots for slot in grp_slots)
                if conflict:
                    continue

                # Agar barcha tekshiruvlardan o'tsa va eng optimali bo'lsa tanlaymiz
                if free > max_free:
                    max_free = free
                    best_teacher = teacher

            # O'qituvchini saqlash
            if best_teacher:
                grp.teacher = best_teacher
                grp.save(update_fields=['teacher'])
                total_assigned_count += 1
            else:
                all_failed_details.append(
                    f"{course.subject.name} ({grp.group_number}-guruh): Mos keladigan konfliktlarsiz o'qituvchi topilmadi")

    if total_assigned_count > 0:
        messages.success(request,
                         f"✅ Jami {total_assigned_count} ta guruhga o'qituvchilar konfliktlarsiz muvaffaqiyatli biriktirildi!")

    if all_failed_details:
        for fail_msg in all_failed_details:
            messages.error(request, f"❌ Joy ajratilmadi: {fail_msg}")
        messages.warning(request,
                         "💡 Ayrim guruhlarga vaqt to'g'ri kelmagani (conflict bergani) sababli o'qituvchi biriktirilmadi. Ularni qo'lda ko'rib chiqishingiz mumkin.")

    return redirect('lesson_list')


@login_required
def group_schedule_debug(request, group_pk):
    grp = get_object_or_404(CourseGroup, pk=group_pk)
    course = grp.course

    import itertools
    start_date = course.start_date
    end_date = course.end_date
    total_lessons = course.total_lessons

    max_wd = 4
    available_wds = list(range(0, max_wd + 1))

    # ── MANTIQLARNI BU YERDA TO'LIQ ANIQHLAYMIZ ──
    if total_lessons >= 20:
        needed_wds = [0, 2, 4]
        candidate_wd_sets = [tuple(needed_wds)]
        days_needed = 3  # <--- Buni belgilash shart
    elif 12 <= total_lessons < 20:
        needed_wds = [1, 3]
        candidate_wd_sets = [tuple(needed_wds)]
        days_needed = 2  # <--- Buni belgilash shart
    else:
        # 8 para: FAQAT bitta kun/hafta (asosiy find_schedule_for_group bilan bir
        # xil mantiq) — guruh raqamiga qarab boshlang'ich kun aylantiriladi,
        # shunda barcha guruhlar bir xil kunga to'planib qolmaydi.
        start_offset = (grp.group_number - 1) % len(available_wds)
        rotated_wds = available_wds[start_offset:] + available_wds[:start_offset]
        candidate_wd_sets = [(wd,) for wd in rotated_wds]
        days_needed = 1  # <--- Buni belgilash shart
    # ─────────────────────────────────────────────

    week_monday = start_date - timedelta(days=start_date.weekday())

    teacher_id = grp.teacher_id
    student_ids = list(grp.students.values_list('id', flat=True))
    student_id_set = set(student_ids)

    # ── Grid (Diagnostika) qismi avvalgidek qoladi ──
    grid = []
    for wd in available_wds:
        d = week_monday + timedelta(days=wd)
        in_range = start_date <= d <= end_date
        blocks = []
        for (p1, p2) in VALID_PARA_PAIRS:
            reasons = []
            if in_range:
                for pi in (p1, p2):
                    if teacher_id:
                        for sc in GroupSchedule.objects.filter(
                                date=d, start_time=PARA_TIMES[pi][0], group__teacher_id=teacher_id,
                        ).exclude(group=grp).select_related('group__course__subject'):
                            reasons.append(
                                f"👨‍🏫 Ustoz band — {sc.group.course.subject} ({sc.group.group_number}-guruh), {PARA_TIMES[pi][0].strftime('%H:%M')}"
                            )
                    if student_ids:
                        for sc in GroupSchedule.objects.filter(
                                date=d, start_time=PARA_TIMES[pi][0],
                                group__students__id__in=student_ids,
                        ).exclude(group=grp).select_related('group__course__subject') \
                                .prefetch_related('group__students').distinct():
                            busy_names = [str(s) for s in sc.group.students.all() if s.id in student_id_set]
                            if busy_names:
                                reasons.append(
                                    f"🎓 Talaba(lar) band: {', '.join(busy_names)} — "
                                    f"{sc.group.course.subject} ({sc.group.group_number}-guruh), "
                                    f"{PARA_TIMES[pi][0].strftime('%H:%M')}"
                                )
            else:
                reasons.append("Kurs muddatidan tashqarida")

            blocks.append({
                'label': f"{PARA_TIMES[p1][0].strftime('%H:%M')}–{PARA_TIMES[p2][1].strftime('%H:%M')}",
                'free': in_range and not reasons,
                'reasons': reasons,
            })
        grid.append({
            'wd': wd, 'wd_name': WEEKDAY_NAMES[wd], 'date': d,
            'in_range': in_range, 'blocks': blocks,
        })

    # ── Qaysi kunlar kombinatsiyasi ishlaydi/ishlamaydi ──
    combo_results = []
    winning_combo = None
    for combo in candidate_wd_sets:
        detail = []
        ok_days = 0
        for wd in combo:
            row = grid[wd]
            # row['blocks'] ichidagi 'free' ni tekshiramiz
            any_free = row['in_range'] and any(b['free'] for b in row['blocks'])
            if any_free:
                ok_days += 1
            detail.append({'wd': wd, 'wd_name': WEEKDAY_NAMES[wd], 'ok': any_free})

        success = ok_days >= days_needed
        combo_results.append({
            'combo_names': [WEEKDAY_NAMES[w] for w in combo],
            'success': success,
            'detail': detail,
        })
        if success and winning_combo is None:
            winning_combo = combo_results[-1]['combo_names']

    context = {
        'grp': grp, 'course': course, 'days_needed': days_needed,
        'week_monday': week_monday, 'grid': grid,
        'combo_results': combo_results, 'winning_combo': winning_combo,
    }
    return render(request, 'raspisaniya/group_schedule_debug.html', context)


@login_required
def teacher_assignment_status(request, course_pk):
    """
    Kurs uchun o'qituvchi taqsimlash holati — JSON.
    Har bir guruh uchun: o'qituvchi biriktirilganmi, kim biriktirilgan.
    """
    course  = get_object_or_404(Course, pk=course_pk)
    groups  = course.groups.select_related('teacher').prefetch_related('students')
    result  = []

    for grp in groups:
        result.append({
            'group_number': grp.group_number,
            'student_count': grp.students.count(),
            'teacher': grp.teacher.first_name if grp.teacher else None,
            'is_scheduled': grp.is_scheduled,
        })

    # Qancha o'qituvchi kerak
    candidates = Teacher.objects.filter(subjects=course.subject).count()
    unassigned = sum(1 for r in result if not r['teacher'])

    return JsonResponse({
        'groups':             result,
        'total_groups':       len(result),
        'unassigned_count':   unassigned,
        'available_teachers': candidates,
        'needs_more':         max(0, unassigned - candidates),
    })



HEADER_FONT = Font(name="Arial", size=10, bold=True, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="1A237E", end_color="1A237E", fill_type="solid")
CENTER_ALIGN = Alignment(horizontal="center", vertical="center")
LEFT_ALIGN = Alignment(horizontal="left", vertical="center")


def format_excel_sheet(ws):
    """Excel jadvalini chiroyli formatlash uchun yordamchi funksiya"""
    ws.row_dimensions[1].height = 28
    for cell in ws[1]:
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER_ALIGN

    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = get_column_letter(col[0].column)   # ✅ to'g'irlandi
        ws.column_dimensions[col_letter].width = max(max_len + 3, 10)
        for cell in col:
            if cell.row > 1:
                cell.alignment = LEFT_ALIGN if cell.column == 1 else CENTER_ALIGN


@login_required
def export_attendance_only_excel(request, group_pk):
    """FAQAT DAVOMATNI EXCELGA YUKLASH"""
    group = get_object_or_404(CourseGroup, pk=group_pk)
    schedules = group.schedule.all().order_by('date', 'lesson_number')
    students = group.students.all().order_by('first_name')
    total_lessons = schedules.count()

    # Barcha davomat yozuvlarini bir martada olib, tez qidirish uchun lug'atga solamiz
    attendance_qs = Attendance.objects.filter(
        schedule__group=group
    ).values('student_id', 'schedule_id', 'is_present')
    att_map = {(a['student_id'], a['schedule_id']): a['is_present'] for a in attendance_qs}

    rows = []
    for student in students:
        cells = []
        came = missed = 0
        for sched in schedules:
            val = att_map.get((student.id, sched.id))
            if val is True:
                came += 1
                cells.append('present')
            elif val is False:
                missed += 1
                cells.append('absent')
            else:
                cells.append('none')

        missed_percent = round(missed / total_lessons * 100) if total_lessons > 0 else 0
        is_blocked = missed_percent > 25 and not group.teacher_can_edit

        rows.append({
            'student': student,
            'cells': cells,
            'came': came,
            'missed': missed,
            'missed_percent': missed_percent,
            'is_blocked': is_blocked,
        })

    wb = Workbook()
    ws = wb.active
    ws.title = "Davomat"

    headers = ["# Talaba"]
    for sched in schedules:
        headers.append(f"{sched.date.strftime('%d.%m')} / {sched.lesson_number}-dars")
    headers.extend(["Keldi", "Kelmadi", "%"])
    ws.append(headers)

    for idx, row in enumerate(rows, start=1):
        name = f"{idx}. {row['student'].first_name}"
        if row['is_blocked']:
            name += " (bloklangan)"
        row_data = [name]
        for cell in row['cells']:
            if cell == 'present':
                row_data.append("✓")
            elif cell == 'absent':
                row_data.append("✗")
            else:
                row_data.append("—")
        row_data.extend([row['came'], row['missed'], f"{row['missed_percent']}%"])
        ws.append(row_data)

    format_excel_sheet(ws)

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename=davomat_{group.group_number}.xlsx'
    wb.save(response)
    return response


@login_required
def export_grades_only_excel(request, group_pk):
    """FAQAT KUNLIK BAHOLARNI EXCELGA YUKLASH"""
    group = get_object_or_404(CourseGroup, pk=group_pk)
    schedules = group.schedule.all().order_by('date', 'lesson_number')
    students = group.students.all().order_by('first_name')
    total_lessons = schedules.count()

    # Har bir darsning JN (30) ichidagi ulushi
    per_lesson_max = (30 / total_lessons) if total_lessons > 0 else 0

    attendance_qs = Attendance.objects.filter(
        schedule__group=group
    ).values('student_id', 'schedule_id', 'is_present')
    att_map = {(a['student_id'], a['schedule_id']): a['is_present'] for a in attendance_qs}

    grade_qs = DailyGrade.objects.filter(
        schedule__group=group
    ).values('student_id', 'schedule_id', 'score')
    grade_map = {(g['student_id'], g['schedule_id']): g['score'] for g in grade_qs}

    rows = []
    for student in students:
        cells = []
        missed = 0
        total_score = 0.0
        for sched in schedules:
            is_present = att_map.get((student.id, sched.id))
            score = grade_map.get((student.id, sched.id))

            if is_present is False:
                missed += 1
                cells.append({'att': 'absent', 'score': None})
            elif score is not None:
                # ✅ TUZATILDI: xom ball emas, 30 balllik tizimga normallashtirilgan qiymat qo'shiladi
                normalized = (score / 100) * per_lesson_max
                total_score += normalized
                cells.append({'att': None, 'score': score})  # jadvalda xom ball ko'rsatiladi
            else:
                cells.append({'att': None, 'score': None})

        missed_percent = round(missed / total_lessons * 100) if total_lessons > 0 else 0
        is_blocked = missed_percent > 25 and not group.teacher_can_edit

        rows.append({
            'student': student,
            'cells': cells,
            'total_score': round(total_score, 2),   # ✅ 2 xonagacha yaxlitlash (masalan 3.38)
            'is_blocked': is_blocked,
        })

    wb = Workbook()
    ws = wb.active
    ws.title = "Kunlik baho"

    headers = ["# Talaba"]
    for sched in schedules:
        headers.append(f"{sched.date.strftime('%d.%m')} / {sched.lesson_number}-dars")
    headers.append("JN (30)")
    ws.append(headers)

    for idx, row in enumerate(rows, start=1):
        name = f"{idx}. {row['student'].first_name}"
        if row['is_blocked']:
            name += " (bloklangan)"
        row_data = [name]
        for cell in row['cells']:
            if cell['att'] == 'absent':
                row_data.append("X")
            elif cell['score'] is not None:
                row_data.append(int(cell['score']))
            else:
                row_data.append("—")
        row_data.append(row['total_score'])
        ws.append(row_data)

    format_excel_sheet(ws)

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename=baholar_{group.group_number}.xlsx'
    wb.save(response)
    return response